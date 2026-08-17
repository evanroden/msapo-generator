"""Employee-reimbursement workbook and submission-packet generation.

The official JDE workbook remains the source document.  This module fills only
its shaded input cells, adds a receipt worksheet after the form, and can ask
LibreOffice/Gotenberg to render the complete workbook as one PDF packet.  It
does not send email or approve an expense report.

WHO DEPENDS ON THIS: app/expense_ui.py drives the whole workflow through it,
and app/receipt_analyzer.py reuses parse_expense_amount and
receipt_preview_bytes so the analyzer and the generator cannot disagree about
what an amount is or what a readable receipt is.

Two properties are worth stating before reading anything below.

FAIL CLOSED, NEVER PARTIALLY. Every write path is preceded by validation
(validate_expense_report) and by _verify_template_anchors, which re-checks
four known strings in the supplied .xlsx. A template whose rows moved would
otherwise be filled at the OLD coordinates and produce a form that looks
complete and codes every line to the wrong place. Nothing here writes to disk
or to a network; the caller receives bytes or an exception.

THE RENDERED PDF IS THE SUBMISSION ARTIFACT, the .xlsx is an editable
convenience. That asymmetry is why build_expense_package treats a renderer
failure as "Excel ready, PDF unavailable" rather than as an error, and why
email_attachments_for_package returns the PDF only.
"""

from __future__ import annotations

import hashlib
import re
import shutil
import subprocess
import tempfile
from collections.abc import Sequence
from copy import copy
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from math import ceil
from pathlib import Path
from urllib.parse import urljoin

import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.pagebreak import Break
from PIL import Image, ImageDraw, ImageFont, ImageOps

from app.config import GOTENBERG_URL, PDF_BACKEND
from app.job_numbers import (
    JOB_NUMBER_OPTIONS,
    job_number_identifier,
    job_numbers_for_contract,
)


BASE_DIR = Path(__file__).resolve().parent.parent
EXPENSE_TEMPLATE_PATH = (
    BASE_DIR
    / "templates"
    / "Employee_Reimbursement_Expense_Report_JDE_10012025.xlsx"
)

# These two strings are a cross-module vocabulary, not private labels: they are
# the values app/receipt_analyzer.py normalizes its section guess onto and the
# values app/expense_ui.py stores in session state. Renaming either strands
# in-progress drafts on a section that matches neither branch below, which is
# an item silently routed into the Miscellaneous block of the official form.
EXPENSE_SECTION_MISC = "miscellaneous"
EXPENSE_SECTION_ENTERTAINMENT = "entertainment"
# Declarative pairing only -- nothing in app/ or tests/ reads EXPENSE_SECTIONS
# today (verified by ripgrep). Kept because it is the one place stating that
# these two are the complete set; do not treat it as a live enum.
EXPENSE_SECTIONS = (EXPENSE_SECTION_MISC, EXPENSE_SECTION_ENTERTAINMENT)

ALLOCATION_JOB = "job"
# ALLOCATION_WORK_ORDER / ALLOCATION_OVERHEAD are NOT dead and must not be
# deleted: they are the values a stale session or an old stored allocation can
# still carry, and tests/test_expense_report.py constructs both to prove that
# allocation_problems and _fill_allocation reject them. Removing them would
# leave the rejection untested and the columns reachable by an unnamed string.
ALLOCATION_WORK_ORDER = "work_order"
ALLOCATION_OVERHEAD = "overhead"
# RRH reimbursement reports use job coding only. Keep the historical constants
# import-compatible, but reject those routes so Work Order and Other Expenses
# columns can never be populated by this workflow.
#
# ALLOCATION_KINDS itself has no reader anywhere in app/ or tests/ (verified by
# ripgrep); the actual gate is the `!= ALLOCATION_JOB` test in
# allocation_problems and again in _fill_allocation. Widening this tuple
# therefore enables NOTHING on its own -- which is the safe direction, but do
# not read it as the switch that turns work orders back on.
ALLOCATION_KINDS = (ALLOCATION_JOB,)

# The official form's own row numbers. These are coordinates in a supplied
# .xlsx, not preferences: MILEAGE_ROWS ends at 17 because row 18 holds
# =SUM(H10:H17), and the same pairing holds for 39 and 59 below. Change a range
# and you must change the matching formula in build_expense_workbook and
# re-verify against the template, or the total silently stops covering the rows
# that were written.
MILEAGE_ROWS = tuple(range(10, 18))
MISCELLANEOUS_ROWS = tuple(range(24, 39))
ENTERTAINMENT_ROWS = tuple(range(45, 59))
# Capacity is derived from the row ranges rather than typed, so a form limit
# and its blocking message can never drift apart. The limits BLOCK; the fill
# loops use zip(), which would otherwise drop the overflow rows in silence.
MAX_MILEAGE_ITEMS = len(MILEAGE_ROWS)
MAX_MISCELLANEOUS_ITEMS = len(MISCELLANEOUS_ROWS)
MAX_ENTERTAINMENT_ITEMS = len(ENTERTAINMENT_ROWS)
# The supplied form states the total must EXCEED $20.00, so the comparison is
# `<=` and $20.00 exactly is refused. Not a rounding cushion -- do not relax it
# to `<` without the product owner.
MINIMUM_REIMBURSEMENT = Decimal("20.00")

# Official IRS business-mileage rates. The IRS made a mid-year change in 2026,
# so rates are selected from the travel date rather than the report date.
#
# The absence of a row for the CURRENT year's successor is deliberate and must
# stay that way. A travel date outside every window resolves to None, and
# validate_expense_report turns that into a blocking message naming the exact
# date. The tempting alternatives -- carry the last known rate forward, or
# extend the final window's end date -- both reimburse an employee at an
# expired rate on a document that goes to accounting, with no indication that a
# rate was assumed rather than published. Add a row only from the published
# IRS notice.
_IRS_BUSINESS_MILEAGE_RATES: tuple[tuple[date, date, Decimal], ...] = (
    (date(2024, 1, 1), date(2024, 12, 31), Decimal("0.67")),
    (date(2025, 1, 1), date(2025, 12, 31), Decimal("0.70")),
    (date(2026, 1, 1), date(2026, 6, 30), Decimal("0.725")),
    (date(2026, 7, 1), date(2026, 12, 31), Decimal("0.76")),
)

_MAX_RECEIPT_PAGES_PER_FILE = 10
# Raster-pixel ceiling checked BEFORE any decode. A few hundred KB of
# compressed input can declare a 20000x20000 page, so validating after the
# raster exists is validating after the damage -- see _validate_receipt_
# dimensions and the matching ordering fix in app/ocr.py.
_MAX_RECEIPT_PIXELS = 40_000_000
# Printed-receipt legibility versus mail size. Receipt pages are rendered into
# the workbook and then into the PDF the approver reads, so this is a
# print-quality decision, not a thumbnail size.
_ATTACHMENT_MAX_SIZE = (1200, 1600)
# One receipt page occupies exactly this many worksheet rows, and the manual
# page break in _build_receipt_sheet is placed on the boundary. See that
# function for the page arithmetic these numbers come from.
_RECEIPT_PAGE_ROWS = 58
# RAW bytes, checked before MIME encoding inflates the attachment by about a
# third. 18 MB raw lands near 24 MB on the wire, which is under the common
# 25 MB Exchange/Outlook ceiling -- so the warning fires while the employee can
# still act on it, rather than after the send bounces. Anyone raising this
# should reason in ENCODED bytes, which is what the mail server measures.
_EMAIL_SAFE_RAW_ATTACHMENT_BYTES = 18 * 1024 * 1024
# Collapses every run of unsafe characters to a single "_". Used only for the
# download/attachment base name; it deliberately drops non-ASCII entirely, so
# expense_report_basename has a fallback for a name that is left empty by it.
_SAFE_FILENAME_RE = re.compile(r"[^A-Za-z0-9._-]+")
# Absolute paths, in preference order, into the CONTAINER's font tree. These
# are a contract with the Dockerfile, not a search list: the fallback was
# fiction until 2026-08-12 because the image installed fonts-dejavu-core, which
# does not ship the italic serif face, so signature rendering depended entirely
# on fonts-urw-base35 and an apt-list trim would have failed the signature step
# closed in production. tests/test_expense_deployment.py now maps each font
# DIRECTORY here to the Debian package providing it and asserts the Dockerfile
# installs it -- add a candidate from a new font family and that test fails
# until the package is added too.
_SIGNATURE_FONT_CANDIDATES = (
    Path("/usr/share/fonts/opentype/urw-base35/Z003-MediumItalic.otf"),
    Path("/usr/share/fonts/truetype/dejavu/DejaVuSerif-Italic.ttf"),
)
# Signature layout. The font is fitted to this canvas at render time rather than
# fixed, so a long name shrinks instead of being clipped.
_SIGNATURE_CANVAS = (1600, 260)
_SIGNATURE_MARGIN = (24, 18)
_SIGNATURE_MAX_FONT_SIZE = 112
# Absolute floor. If a name still does not fit at this size the renderer raises
# rather than emitting a clipped signature — see employee_signature_png.
_SIGNATURE_MIN_FONT_SIZE = 12
_SIGNATURE_MAX_NAME_CHARS = 160


class ExpenseReportError(ValueError):
    """Raised when an expense package cannot be generated safely."""


@dataclass(frozen=True)
class ExpenseAllocation:
    """One JDE allocation route and its route-specific coding fields.

    Only `kind == ALLOCATION_JOB` is accepted by this workflow, and only
    job_number / account_cost_type / cost_code_or_wo_type are ever written to
    the form (columns I, J, K).

    The remaining fields are NOT dead. app/memory.py's expense-profile table
    has a column per field and app/expense_ui.py reads and writes them by name,
    so they exist to keep that schema round-trippable across the versions of
    the app that did populate them. Deleting one breaks a stored-profile read;
    populating one does not put it on the form, because _fill_allocation
    writes three columns and refuses any non-job kind.
    """

    kind: str
    job_number: str = ""
    service_center: str = ""
    account_cost_type: str = ""
    cost_code_or_wo_type: str = ""
    work_order_number: str = ""
    company_number: str = ""
    department_number: str = ""
    ou_number: str = ""
    gl_account_number: str = ""


@dataclass(frozen=True)
class ExpenseItem:
    """One reviewed reimbursement line from an uploaded receipt.

    A LINE, not a receipt. One uploaded receipt split across several business
    purposes or coding routes produces several ExpenseItems that share
    source_receipt_id and file_bytes, and the whole grouping machinery below
    (_receipt_source_id, _receipt_groups, _unique_receipt_count) exists so the
    source image is attached exactly ONCE while each line still occupies its
    own form row.

    `amount` is intentionally typed loosely: the UI hands over the raw string
    the employee typed, and parse_expense_amount is the single place that
    decides whether it is money. Never do arithmetic on this field directly.
    """

    receipt_id: str
    filename: str
    file_bytes: bytes
    transaction_date: date | None
    description: str
    amount: Decimal | str | None
    section: str
    allocation: ExpenseAllocation
    merchant_name: str = ""
    contact_name: str = ""
    # Split receipts use one unique receipt_id per reimbursement line while
    # sharing source_receipt_id. The source image is attached only once.
    source_receipt_id: str = ""


@dataclass(frozen=True)
class MileageItem:
    """One business-mileage row on the official reimbursement form."""

    entry_id: str
    transaction_date: date | None
    purpose: str
    destination: str
    miles: Decimal | str | float | None
    allocation: ExpenseAllocation


@dataclass(frozen=True)
class ExpenseReportDetails:
    """Employee, routing, and approval information for one report."""

    account: str
    employee_name: str
    employee_number: str
    employee_home_bu: str
    report_date: date
    approver_name: str
    approver_email: str
    mail_destination: str = "home"
    satellite_office: str = ""
    employee_signature_confirmed: bool = False


@dataclass(frozen=True)
class ExpensePackage:
    """Generated workbook plus an optional rendered PDF packet.

    `pdf_bytes is None` with a non-empty `pdf_error` is a NORMAL, supported
    outcome, not a broken package: the workbook is complete and validated, the
    renderer simply was not available. Callers must branch on pdf_bytes and
    keep offering the workbook. Treating this whole object as failed is the
    exact defect that discarded finished reports before 2026-08-12.

    `total` is computed in Python from the reviewed items; the workbook shows
    its own =H18+H39+H59. See mileage_reimbursement for the one rounding case
    in which those two can differ by a cent.
    """

    basename: str
    workbook_bytes: bytes
    pdf_bytes: bytes | None
    total: Decimal
    receipt_count: int
    pdf_error: str = ""
    mileage_count: int = 0


def parse_expense_amount(value: object) -> Decimal | None:
    """Return a positive two-decimal amount, or ``None`` for invalid input.

    The SINGLE currency parser for this app -- app/receipt_analyzer.py and
    app/expense_ui.py both route through it, so the amount the analyzer
    displays and the amount the workbook writes cannot disagree.

    Strictness is the feature. The regex requires the whole remaining string to
    be a number, which is why "amount 12.00" is rejected rather than salvaged
    (tests/test_expense_report.py pins that). A lenient "find the digits"
    parser would accept a model's prose, a phone number, or an invoice
    reference and put it on a reimbursement form as a dollar figure.

    Rejects zero and negatives, so callers can use `or Decimal("0")` freely and
    a refund/credit line can never subtract from a total silently. Rejects
    anything above ~$10M, which is a typo guard, not a policy limit.

    Note the asymmetry between input types: strings go through the regex,
    while Decimal/int/float bypass it and are only range-checked and quantized.
    A float therefore rounds with Decimal's default ROUND_HALF_EVEN, unlike the
    ROUND_HALF_UP the UI uses when it allocates a receipt across items. In
    practice everything user-entered arrives as a string with at most two
    decimals, so this does not bite -- but do not start feeding it raw floats.
    """
    if isinstance(value, Decimal):
        amount = value
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        amount = Decimal(str(value))
    else:
        text = str(value or "").strip()
        if not text:
            return None
        # Strip only the three decorations a person or a receipt actually
        # types: a "USD" suffix, a leading "$", and accounting parentheses.
        # Parentheses are normalized to a minus sign and then REJECTED by the
        # positivity check below -- the conversion exists so "(31.25)" fails
        # loudly as a negative rather than parsing as a positive $31.25.
        text = re.sub(r"\bUSD\b", "", text, flags=re.IGNORECASE).strip()
        if text.startswith("$"):
            text = text[1:].strip()
        if text.startswith("(") and text.endswith(")"):
            text = f"-{text[1:-1]}"
        # fullmatch, not search: the whole remainder must be the number. The
        # alternation forbids ragged grouping ("1,23,4") while accepting both
        # "1234.50" and "1,234.50", and at most two decimals so no rounding
        # decision is ever made on user text.
        if not re.fullmatch(
            r"-?(?:\d+|\d{1,3}(?:,\d{3})+)(?:\.\d{1,2})?",
            text,
        ):
            return None
        try:
            amount = Decimal(text.replace(",", ""))
        except InvalidOperation:
            return None
    if not amount.is_finite() or amount <= 0 or amount > Decimal("9999999.99"):
        return None
    return amount.quantize(Decimal("0.01"))


def parse_mileage(value: object) -> Decimal | None:
    """Return a positive mileage quantity, or ``None`` for invalid input.

    Same shape as parse_expense_amount but deliberately NOT the same function:
    miles are not currency, so no "$", no "USD" and no accounting parentheses
    are accepted here. A value like "$40" in the miles box is a field mix-up
    and must fail rather than be cleaned up into 40 miles.

    The 100000 ceiling is a typo guard on a hand-typed odometer figure.
    """
    if isinstance(value, Decimal):
        miles = value
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        miles = Decimal(str(value))
    else:
        text = str(value or "").strip()
        if not re.fullmatch(r"(?:\d+|\d{1,3}(?:,\d{3})+)(?:\.\d{1,2})?", text):
            return None
        try:
            miles = Decimal(text.replace(",", ""))
        except InvalidOperation:
            return None
    if not miles.is_finite() or miles <= 0 or miles > Decimal("100000"):
        return None
    return miles.quantize(Decimal("0.01"))


def irs_business_mileage_rate(travel_date: date | None) -> Decimal | None:
    """Return the configured IRS business rate for one travel date.

    TRAVEL date, never the report date. The IRS changed the rate mid-2026, so a
    July report covering June travel must pay the June rate; keying off the
    report date silently overpays or underpays every row that crosses the
    boundary. tests/test_expense_report.py pins both sides of 2026-06-30.

    Returns None for an unconfigured date, which callers must surface as a
    blocker rather than substitute a rate for -- see the table above.
    """
    if travel_date is None:
        return None
    for effective, through, rate in _IRS_BUSINESS_MILEAGE_RATES:
        if effective <= travel_date <= through:
            return rate
    return None


def mileage_reimbursement(item: MileageItem) -> Decimal | None:
    """Return the rounded reimbursement for one valid mileage row.

    None when the row is not yet valid (unreadable miles or an unconfigured
    rate); callers treat that as zero for display and validation blocks it
    before generation.

    KNOWN DIVERGENCE, do not "simplify" either side into the other: this
    quantize() uses Decimal's default ROUND_HALF_EVEN, while the cell this row
    also receives is the live formula =ROUND(G*rate,2), and Excel's ROUND is
    half-UP. A product landing exactly on a half-cent -- 1.15 miles at $0.70 is
    0.8050 -- gives 0.80 here and 0.81 in the workbook, so the total quoted in
    the approval email can trail the form's Q60 by a cent. The formula stays in
    the cell because the form must remain an editable workbook whose totals
    recompute when the employee corrects a row.
    """
    miles = parse_mileage(item.miles)
    rate = irs_business_mileage_rate(item.transaction_date)
    if miles is None or rate is None:
        return None
    return (miles * rate).quantize(Decimal("0.01"))


def allocation_problems(
    allocation: ExpenseAllocation,
    *,
    prefix: str = "",
    allowed_job_numbers: Sequence[str] | None = None,
) -> list[str]:
    """Validate the job coding used by RRH reimbursement reports.

    Returns a list of operator-facing phrases, each prefixed with ``prefix``
    so the report-wide message names the row it belongs to. Empty list means
    the coding is complete, NOT that it is correct for the work -- only a human
    knows that.

    ``allowed_job_numbers=None`` means "any catalog job", which is what a
    direct call or a test gets. The UI passes the account's list, so a job
    number left over from a previous account selection is rejected instead of
    being written to a form for a different customer.

    Returning EARLY on a non-job kind is intentional: once the route is wrong
    there is nothing useful to say about its individual fields, and listing
    four more problems buries the one that matters.
    """
    label = f"{prefix}: " if prefix else ""
    problems: list[str] = []
    if allocation.kind != ALLOCATION_JOB:
        return [f"{label}use job coding; work orders and Other Expenses are not used"]
    if allocation.job_number not in JOB_NUMBER_OPTIONS:
        problems.append(f"{label}choose the job number")
    elif (
        allowed_job_numbers is not None
        and allocation.job_number not in allowed_job_numbers
    ):
        problems.append(f"{label}choose a job number for the selected account")
    if not allocation.account_cost_type.strip():
        problems.append(f"{label}enter the account / cost type")
    if not allocation.cost_code_or_wo_type.strip():
        problems.append(f"{label}enter the cost code")
    return problems


def validate_expense_report(
    details: ExpenseReportDetails,
    items: list[ExpenseItem],
    mileage_items: Sequence[MileageItem] = (),
) -> list[str]:
    """Return actionable blocking problems in interaction order.

    THE authoritative gate. build_expense_workbook calls this first and refuses
    to write anything if it returns a non-empty list, which is what lets every
    _fill_* helper below assert its preconditions instead of re-checking them.
    app/expense_ui.py's per-card warnings are a display convenience layered on
    top; an empty list from those never means the report is generatable.

    Empty list means generatable. Order is the order the operator meets the
    fields on the page, not severity, so the first message is the first thing
    to go fix. Messages are de-duplicated at the end because several rows
    legitimately raise the identical phrase.

    NUMBERING CAVEAT: "Receipt N" here counts reimbursement LINES in the order
    supplied, so a receipt split into two lines occupies N and N+1. The
    per-card numbering in the UI and the "Receipt N of M" headers in the
    RECEIPTS worksheet both count unique SOURCES instead. They agree for the
    common unsplit case and diverge once a receipt is split.
    """
    problems: list[str] = []
    for value, field in (
        (details.account, "choose the account / contract"),
        (details.employee_name, "enter the employee name"),
        (details.employee_number, "enter the employee number"),
        (
            details.employee_home_bu,
            "enter the Employee Home Business Unit",
        ),
        (details.approver_name, "enter the contract administrator's name"),
        (details.approver_email, "enter the contract administrator's email"),
    ):
        if not str(value or "").strip():
            problems.append(field)
    if not _looks_like_email(details.approver_email):
        problems.append("enter a valid contract administrator email")
    if details.mail_destination not in {"home", "satellite"}:
        problems.append("choose where the reimbursement check should be mailed")
    if details.mail_destination == "satellite" and not details.satellite_office.strip():
        problems.append("enter the satellite office")
    if not details.employee_signature_confirmed:
        problems.append("confirm the generated employee signature")

    # A mileage-only report is valid, so this is "neither", not "no receipts".
    # Returning early keeps the $20 minimum from firing on an empty report and
    # burying the one message that actually applies.
    if not items and not mileage_items:
        problems.append("include at least one receipt or mileage entry")
        return _deduplicate(problems)

    seen_lines: set[str] = set()
    # source id -> (filename, sha256 of the bytes). Split lines are supposed to
    # be several views of ONE upload; if two lines claim the same source but
    # carry different files, one of them attaches and the other silently does
    # not, so the packet would show a receipt that does not support the row
    # next to it. Hashing rather than keeping the bytes avoids a second full
    # copy of every receipt in memory.
    source_fingerprints: dict[str, tuple[str, bytes]] = {}
    allowed_job_numbers = job_numbers_for_contract(details.account)
    misc_count = 0
    entertainment_count = 0
    total = Decimal("0")
    for index, item in enumerate(items, 1):
        prefix = f"Receipt {index}"
        if not item.receipt_id or item.receipt_id in seen_lines:
            problems.append(f"{prefix}: remove the duplicate reimbursement line")
        seen_lines.add(item.receipt_id)
        source_id = _receipt_source_id(item)
        fingerprint = (item.filename, hashlib.sha256(item.file_bytes).digest())
        if source_id in source_fingerprints:
            if source_fingerprints[source_id] != fingerprint:
                problems.append(
                    f"{prefix}: split lines must use the same uploaded receipt"
                )
        else:
            source_fingerprints[source_id] = fingerprint
        if item.transaction_date is None:
            problems.append(f"{prefix}: enter the transaction date")
        if not item.description.strip():
            problems.append(f"{prefix}: enter the description or business purpose")
        amount = parse_expense_amount(item.amount)
        if amount is None:
            problems.append(f"{prefix}: enter a valid reimbursable amount")
        else:
            total += amount
        if item.section == EXPENSE_SECTION_MISC:
            misc_count += 1
        elif item.section == EXPENSE_SECTION_ENTERTAINMENT:
            entertainment_count += 1
            if not item.contact_name.strip():
                problems.append(f"{prefix}: enter the entertainment contact name")
        else:
            problems.append(f"{prefix}: choose Miscellaneous or Entertainment")
        problems.extend(
            allocation_problems(
                item.allocation,
                prefix=prefix,
                allowed_job_numbers=allowed_job_numbers,
            )
        )
        if not item.file_bytes:
            problems.append(f"{prefix}: upload the receipt image or PDF again")

    seen_mileage: set[str] = set()
    for index, item in enumerate(mileage_items, 1):
        prefix = f"Mileage {index}"
        if not item.entry_id or item.entry_id in seen_mileage:
            problems.append(f"{prefix}: remove the duplicate mileage entry")
        seen_mileage.add(item.entry_id)
        if item.transaction_date is None:
            problems.append(f"{prefix}: enter the travel date")
        if not item.purpose.strip():
            problems.append(f"{prefix}: enter the business purpose")
        if not item.destination.strip():
            problems.append(f"{prefix}: enter the destination")
        if parse_mileage(item.miles) is None:
            problems.append(f"{prefix}: enter valid business miles")
        rate = irs_business_mileage_rate(item.transaction_date)
        if item.transaction_date is not None and rate is None:
            problems.append(
                f"{prefix}: the IRS business-mileage rate for "
                f"{item.transaction_date:%Y-%m-%d} is not configured"
            )
        reimbursement = mileage_reimbursement(item)
        if reimbursement is not None:
            total += reimbursement
        problems.extend(
            allocation_problems(
                item.allocation,
                prefix=prefix,
                allowed_job_numbers=allowed_job_numbers,
            )
        )

    # These three limits are the ONLY thing standing between an over-capacity
    # report and silent data loss: build_expense_workbook fills rows with
    # zip(ROWS, items), which stops at the shorter side without a word. An
    # employee who submitted 16 Miscellaneous receipts would receive a form
    # showing 15 and a total that matched it. Counting per section, because the
    # form's two blocks have different capacities.
    if len(mileage_items) > MAX_MILEAGE_ITEMS:
        problems.append(
            f"split the report: the official form allows {MAX_MILEAGE_ITEMS} "
            "mileage entries"
        )
    if misc_count > MAX_MISCELLANEOUS_ITEMS:
        problems.append(
            f"split the report: the official form allows {MAX_MISCELLANEOUS_ITEMS} "
            "Miscellaneous receipts"
        )
    if entertainment_count > MAX_ENTERTAINMENT_ITEMS:
        problems.append(
            f"split the report: the official form allows {MAX_ENTERTAINMENT_ITEMS} "
            "Entertainment receipts"
        )
    if total <= MINIMUM_REIMBURSEMENT:
        problems.append("the total reimbursement must exceed $20.00")
    return _deduplicate(problems)


def expense_report_total(items: list[ExpenseItem]) -> Decimal:
    """Return the reviewed receipt total; invalid values contribute zero.

    Zero-for-invalid is safe ONLY because validate_expense_report blocks an
    unparseable amount before anything is generated. This runs while the
    operator is still typing, where a half-typed "3." must show as a smaller
    running total rather than raise and blank the page.
    """
    return sum(
        (parse_expense_amount(item.amount) or Decimal("0") for item in items),
        Decimal("0"),
    )


def total_reimbursement(
    items: Sequence[ExpenseItem],
    mileage_items: Sequence[MileageItem] = (),
) -> Decimal:
    """Return the receipt plus mileage reimbursement total.

    The figure shown in the UI summary and quoted in the approval email. Same
    invalid-is-zero contract as expense_report_total; validation is what makes
    the generated number trustworthy.
    """
    receipt_total = expense_report_total(list(items))
    mileage_total = sum(
        (mileage_reimbursement(item) or Decimal("0") for item in mileage_items),
        Decimal("0"),
    )
    return receipt_total + mileage_total


def expense_report_warnings(
    details: ExpenseReportDetails,
    items: list[ExpenseItem],
    mileage_items: Sequence[MileageItem] = (),
) -> list[str]:
    """Return non-blocking checks that require human confirmation.

    Deliberately SEPARATE from validate_expense_report: everything here has a
    legitimate explanation, so none of it may block. Two people buying the same
    coffee at the same shop for the same price on the same day is real, and
    auto-deleting the "duplicate" would delete a genuine expense.

    Numbering counts unique SOURCE receipts in supplied order, which is what
    the UI cards show. It is not the same numbering validate_expense_report
    uses; see the caveat there.
    """
    warnings: list[str] = []
    possible_duplicates: dict[
        tuple[date, Decimal, str], list[tuple[str, int]]
    ] = {}
    source_numbers: dict[str, int] = {}
    date_checked: set[str] = set()
    for item in items:
        source_id = _receipt_source_id(item)
        source_number = source_numbers.setdefault(source_id, len(source_numbers) + 1)
        # date_checked keeps the age checks per SOURCE, not per line. Without
        # it a receipt split into four lines produces the same "dated after the
        # report date" sentence four times, which reads as four bad receipts.
        if item.transaction_date and source_id not in date_checked:
            if item.transaction_date > details.report_date:
                warnings.append(
                    f"Receipt {source_number} is dated after the report date"
                )
            elif (details.report_date - item.transaction_date).days > 366:
                warnings.append(
                    f"Receipt {source_number} is more than one year older than the report date"
                )
            date_checked.add(source_id)
        amount = parse_expense_amount(item.amount)
        # Squashing case and punctuation makes "Dunkin'" and "DUNKIN" the same
        # merchant, which is the point: the duplicate this catches is usually
        # the SAME receipt photographed twice, and the second read rarely
        # spells the merchant identically. An empty merchant disables the
        # check for that line rather than matching every other unnamed one.
        merchant = re.sub(
            r"[^a-z0-9]+", "", item.merchant_name.casefold()
        )
        if item.transaction_date and amount is not None and merchant:
            key = (item.transaction_date, amount, merchant)
            possible_duplicates.setdefault(key, []).append(
                (source_id, source_number)
            )
    for matches in possible_duplicates.values():
        # dict() collapses the matches by source id, so the warning fires only
        # when two DIFFERENT uploads collide. Split lines of one receipt share
        # a source id and legitimately share a date and merchant, so comparing
        # raw list length instead would warn about every split receipt --
        # tests/test_expense_report.py pins that they must not.
        distinct_sources = dict(matches)
        if len(distinct_sources) > 1:
            labels = ", ".join(
                str(number) for number in distinct_sources.values()
            )
            warnings.append(
                f"Receipts {labels} have the same merchant, date, and amount; "
                "confirm they are separate purchases"
            )
    for index, item in enumerate(mileage_items, 1):
        if item.transaction_date and item.transaction_date > details.report_date:
            warnings.append(f"Mileage {index} is dated after the report date")
    return _deduplicate(warnings)


def expense_report_signature(
    details: ExpenseReportDetails,
    items: list[ExpenseItem],
    mileage_items: Sequence[MileageItem] = (),
) -> str:
    """Stable signature used to suppress stale downloads after an edit.

    Covers EVERY reviewed input, including the receipt bytes. app/expense_ui.py
    compares it again at render time, so an edit made after generation
    withholds the previously generated files instead of handing the approver a
    PDF built from the old values. Narrowing this to the "important" fields
    silently reopens that -- a changed receipt image, a corrected contact name
    or a re-coded allocation are all things an approver would be shown
    incorrectly.

    Depends on repr() being stable for these dataclasses, which is why the item
    tuple is spelled out field by field rather than repr(item): file_bytes must
    stay out of that repr and be hashed separately, so a 10 MB receipt is not
    turned into a multi-megabyte string on every rerun.
    """
    digest = hashlib.sha256()
    digest.update(repr(details).encode("utf-8"))
    for item in items:
        digest.update(
            repr(
                (
                    item.receipt_id,
                    item.filename,
                    item.transaction_date,
                    item.description,
                    item.amount,
                    item.section,
                    item.allocation,
                    item.merchant_name,
                    item.contact_name,
                    item.source_receipt_id,
                )
            ).encode("utf-8")
        )
        digest.update(hashlib.sha256(item.file_bytes).digest())
    for item in mileage_items:
        digest.update(repr(item).encode("utf-8"))
    return digest.hexdigest()


def build_expense_workbook(
    details: ExpenseReportDetails,
    items: list[ExpenseItem],
    *,
    mileage_items: Sequence[MileageItem] = (),
    template_path: Path = EXPENSE_TEMPLATE_PATH,
) -> bytes:
    """Fill the official workbook and append printable receipt images.

    Returns .xlsx bytes. Raises ExpenseReportError -- with every blocking
    problem joined into one message -- before touching the template, so a
    rejected report leaves no partially written artifact anywhere.

    ``template_path`` is a parameter only so tests can point at a fixture. The
    production value is the supplied official form and must not be swapped for
    a regenerated lookalike; _verify_template_anchors is what catches that.

    The employee's ORIGINAL receipt bytes are never modified. Everything below
    works on normalized copies.
    """
    problems = validate_expense_report(details, items, mileage_items)
    if problems:
        raise ExpenseReportError("; ".join(problems))
    if not template_path.is_file():
        raise ExpenseReportError(
            "The official employee-reimbursement workbook template is missing."
        )

    workbook = load_workbook(template_path)
    if "EXPENSE REIMBURSEMENT" not in workbook.sheetnames:
        raise ExpenseReportError("The reimbursement template has an unexpected layout.")
    form = workbook["EXPENSE REIMBURSEMENT"]
    # Before ANY write. A template whose rows shifted would otherwise be filled
    # at the old coordinates and produce a complete-looking form with every
    # value one section out of place.
    _verify_template_anchors(form)
    # Collects the BytesIO backing each embedded image so every stream stays
    # referenced until workbook.save() has serialized the drawings below. The
    # list is never read; that is intended, and it is why both helpers take it
    # as a parameter rather than owning their own buffers.
    image_buffers: list[BytesIO] = []
    _fill_report_header(form, details, mileage_items, image_buffers)

    # ORDER MATTERS: _fill_report_header above pre-fills all eight mileage rows
    # with G=0 and a rate formula so unused rows total zero. Doing it after
    # this loop would overwrite every real mileage figure with 0 -- the
    # workbook would still open, still recalculate, and still look finished,
    # with the mileage half of the reimbursement quietly gone.
    ordered_mileage = _mileage_grouped_for_form(mileage_items)
    for row, item in zip(MILEAGE_ROWS, ordered_mileage):
        _fill_mileage_row(form, row, item)

    ordered_items = _items_grouped_for_form(items)
    miscellaneous = [
        item for item in ordered_items if item.section == EXPENSE_SECTION_MISC
    ]
    entertainment = [
        item for item in ordered_items
        if item.section == EXPENSE_SECTION_ENTERTAINMENT
    ]
    # zip() truncates in silence if either list outruns its row range. That is
    # only safe because validate_expense_report already blocked on the same
    # per-section counts above; relaxing those limits without changing this
    # turns an over-capacity report into a quietly short one.
    for row, item in zip(MISCELLANEOUS_ROWS, miscellaneous):
        _fill_expense_row(form, row, item, entertainment=False)
    for row, item in zip(ENTERTAINMENT_ROWS, entertainment):
        _fill_expense_row(form, row, item, entertainment=True)

    # Keep formulas intact for an editable workbook and force Excel/LibreOffice
    # to refresh their cached values when the file opens or renders.
    #
    # Writing computed NUMBERS here instead is the obvious simplification and
    # is wrong twice over. The .xlsx is handed to the employee to correct a row
    # in, and a hardcoded total would then contradict the rows above it. And
    # openpyxl does not evaluate formulas, so the file it saves carries the
    # template's STALE cached results; without fullCalcOnLoad the PDF renderer
    # prints those old numbers -- a wrong total on the approver's copy, with
    # nothing on screen indicating a computation was skipped.
    form["H18"] = "=SUM(H10:H17)"
    form["H39"] = "=SUM(H24:H38)"
    form["H59"] = "=SUM(H45:H58)"
    form["Q60"] = "=H18+H39+H59"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    # Rebuild from scratch rather than append. The template ships without this
    # sheet, but a stale one from an earlier generation must never survive into
    # a regenerated packet -- that is a previous report's receipts attached to
    # this report's form.
    if "RECEIPTS" in workbook.sheetnames:
        workbook.remove(workbook["RECEIPTS"])
    # No sheet at all for a mileage-only report; an empty "RECEIPTS" tab reads
    # as missing receipts. tests/test_expense_report.py pins the sheet list.
    if ordered_items:
        receipt_sheet = workbook.create_sheet("RECEIPTS")
        # Receipt pages follow the same grouped order as the form rows, so the
        # packet remains easy to audit when a report uses multiple coding strings.
        _build_receipt_sheet(receipt_sheet, ordered_items, image_buffers)

    payload = BytesIO()
    workbook.save(payload)
    return payload.getvalue()


def build_expense_package(
    details: ExpenseReportDetails,
    items: list[ExpenseItem],
    *,
    mileage_items: Sequence[MileageItem] = (),
    include_pdf: bool = True,
) -> ExpensePackage:
    """Generate the workbook and, when available, its combined PDF packet.

    The only failure that propagates is a rejected report:
    build_expense_workbook raises ExpenseReportError before the try below, so
    an invalid report never reaches the renderer. A renderer failure does NOT
    propagate -- see the comment inside the except.

    ``include_pdf=False`` exists for tests and for callers that only need the
    workbook; it is not a fallback for a broken renderer, because the resulting
    package looks identical to a renderer failure with no pdf_error to show.
    """
    workbook_bytes = build_expense_workbook(
        details,
        items,
        mileage_items=mileage_items,
    )
    pdf_bytes = None
    pdf_error = ""
    if include_pdf:
        try:
            pdf_bytes = convert_expense_workbook_to_pdf(workbook_bytes)
        except Exception as exc:  # noqa: BLE001 - see below; this must not re-raise
            # Preserve the editable workbook so a renderer outage never forces
            # re-entry, but the UI withholds the approval draft because RRH's
            # approved email submission artifact is the PDF.
            #
            # This deliberately catches Exception rather than ExpenseReportError.
            # The workbook above is already built and validated; the ONLY thing
            # that can fail here is the renderer, and every renderer failure must
            # degrade to "Excel ready, PDF unavailable" rather than discarding
            # work the operator already typed. ExpenseReportError alone was not
            # sufficient: subprocess.run(..., timeout=120) raises
            # subprocess.TimeoutExpired (a receipt-heavy workbook is exactly the
            # case that exceeds the budget), and the Gotenberg backend can raise
            # requests.ConnectionError before any wrapping runs. Both escaped,
            # the caller's generic handler dropped the package, and the
            # Excel-only fallback this block exists to enable became unreachable.
            pdf_error = str(exc) or exc.__class__.__name__
    basename = expense_report_basename(details)
    return ExpensePackage(
        basename=basename,
        workbook_bytes=workbook_bytes,
        pdf_bytes=pdf_bytes,
        total=total_reimbursement(items, mileage_items),
        receipt_count=_unique_receipt_count(items),
        pdf_error=pdf_error,
        mileage_count=len(mileage_items),
    )


def convert_expense_workbook_to_pdf(workbook_bytes: bytes) -> bytes:
    """Render both workbook sheets as one PDF, using the configured backend.

    Returns PDF bytes. Raises ExpenseReportError on every reachable failure --
    unknown backend, missing binary, non-200 service response, absent or
    non-PDF output. Callers must assume this can also raise things that are NOT
    ExpenseReportError (subprocess.TimeoutExpired, requests transport errors);
    build_expense_package catches Exception for exactly that reason.

    An UNRECOGNIZED PDF_BACKEND raises rather than falling through to a
    default. Silently substituting LibreOffice for a misspelled "gotenburg"
    would make a deployment misconfiguration invisible on the one machine where
    both happen to work.
    """
    backend = str(PDF_BACKEND or "libreoffice").strip().lower()
    if backend == "gotenberg":
        endpoint = urljoin(GOTENBERG_URL.rstrip("/") + "/", "forms/libreoffice/convert")
        response = requests.post(
            endpoint,
            files={
                "files": (
                    "expense_report.xlsx",
                    workbook_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            timeout=120,
        )
        # Sniff the magic bytes as well as the status. A proxy, a captive
        # portal or a misrouted host answers 200 with an HTML error page, and
        # that page would otherwise be attached to the approval email as the
        # employee's expense report.
        if response.status_code != 200 or not response.content.startswith(b"%PDF"):
            raise ExpenseReportError(
                f"The PDF service could not render the expense report (HTTP "
                f"{response.status_code})."
            )
        return response.content
    if backend != "libreoffice":
        raise ExpenseReportError(
            f"Expense-report PDF generation does not support PDF_BACKEND={backend!r}."
        )

    executable = shutil.which("libreoffice") or shutil.which("soffice")
    if not executable:
        raise ExpenseReportError(
            "LibreOffice is unavailable, so the Excel report was created but the "
            "combined PDF could not be rendered."
        )
    with tempfile.TemporaryDirectory(prefix="expense-report-") as temp_name:
        temp = Path(temp_name)
        source = temp / "expense_report.xlsx"
        source.write_bytes(workbook_bytes)
        # A PRIVATE LibreOffice profile per conversion, inside the temp dir
        # that is about to be deleted. Concurrent conversions sharing the
        # default ~/.config profile contend on its lock, and the loser exits
        # successfully having written no file at all -- which surfaces here as
        # the "could not render" branch below, intermittently, only under load.
        # The Writer path in app/pdf_converter.py had to learn this separately.
        profile_uri = (temp / "libreoffice-profile").resolve().as_uri()
        result = subprocess.run(
            [
                executable,
                "--headless",
                f"-env:UserInstallation={profile_uri}",
                "--convert-to",
                "pdf:calc_pdf_Export",
                "--outdir",
                str(temp),
                str(source),
            ],
            capture_output=True,
            text=True,
            timeout=120,
            check=False,
        )
        # check=False, then test for the OUTPUT FILE. LibreOffice routinely
        # exits 0 while converting nothing (missing import filter, profile
        # lock, unreadable input), so trusting the return code is how a
        # renderer failure becomes a silent one. The file's existence is the
        # only honest success signal, and its first four bytes are the second.
        rendered = temp / "expense_report.pdf"
        if not rendered.is_file():
            detail = (result.stderr or result.stdout or "unknown error").strip()[:300]
            raise ExpenseReportError(
                f"LibreOffice could not render the combined PDF: {detail}"
            )
        payload = rendered.read_bytes()
        if not payload.startswith(b"%PDF"):
            raise ExpenseReportError("LibreOffice returned an invalid PDF file.")
        return payload


def expense_report_basename(details: ExpenseReportDetails) -> str:
    """Return a filesystem-safe, useful attachment base name.

    No extension; callers append ".xlsx"/".pdf" so both artifacts of one
    report share a stem and sort together in the approver's inbox.

    Never returns an empty or unsafe stem: a name that sanitizes to nothing --
    a name written entirely in a non-Latin script, for instance -- falls back
    to "Employee" rather than producing a file called ".pdf". The 120-character
    cap keeps the eventual path inside conservative filesystem limits.
    """
    employee = _SAFE_FILENAME_RE.sub("_", details.employee_name.strip()).strip("_ .")
    employee = employee or "Employee"
    return f"{details.report_date:%Y.%m.%d}_Expense_Report_{employee}"[:120]


def email_attachments_for_package(package: ExpensePackage) -> list[tuple[str, bytes]]:
    """Return the submission PDF; Excel remains a separate editable download.

    PDF ONLY. Adding package.workbook_bytes here would attach the editable
    workbook to the approval email, which the 2026-08-11 handoff forbids
    (invariants 1 and 2): the approved submission artifact is the rendered PDF,
    and an approver who receives a spreadsheet can silently alter it.

    Returns [] when there is no PDF, which is a real state -- see
    ExpensePackage. Callers must NOT build a draft from an empty list; an .eml
    that claims an attachment and carries none is worse than no draft at all,
    so app/expense_ui.py raises instead.
    """
    if not package.pdf_bytes:
        return []
    return [(f"{package.basename}.pdf", package.pdf_bytes)]


def email_attachment_size_warning(package: ExpensePackage) -> str:
    """Warn when the submission PDF may exceed a common mail-server limit.

    Returns "" when there is nothing to say. Advisory ONLY: it never removes,
    downsamples or substitutes the attachment. A receipt-heavy report is still
    a valid report, and silently swapping in a smaller artifact would hand the
    approver a different document than the employee reviewed.
    """
    if not package.pdf_bytes:
        return ""
    size = len(package.pdf_bytes)
    if size <= _EMAIL_SAFE_RAW_ATTACHMENT_BYTES:
        return ""
    return (
        f"The PDF attachment is {size / (1024 * 1024):.1f} MB before email "
        "encoding. Outlook or the mail server may require a smaller report."
    )


def receipt_preview_bytes(file_bytes: bytes, filename: str) -> bytes:
    """Return a bounded JPEG preview of the first receipt page/frame.

    Doubles as the upload PREFLIGHT, which is why app/receipt_analyzer.py calls
    it and throws the result away for PDFs: rendering one page still validates
    the whole source's page count and every page's raster size, so an
    unacceptable receipt is rejected locally instead of after an upload to the
    vision API. Raises ExpenseReportError with the operator-facing reason.

    For images this IS the payload the analyzer sends -- it applies EXIF
    rotation and bounds a phone photo, both of which the raw upload lacks.
    """
    pages = receipt_attachment_pages(file_bytes, filename, render_limit=1)
    if not pages:
        raise ExpenseReportError("The receipt did not contain a readable page.")
    return pages[0]


def receipt_attachment_pages(
    file_bytes: bytes,
    filename: str,
    *,
    max_pages: int = _MAX_RECEIPT_PAGES_PER_FILE,
    render_limit: int | None = None,
) -> list[bytes]:
    """Validate and normalize receipt pages to compact, print-readable JPEGs.

    ``max_pages`` is the security/business limit for the source. ``render_limit``
    only limits returned pages, allowing a first-page preview without weakening
    validation of a multipage source.

    Returns at least one JPEG or raises ExpenseReportError -- never an empty
    list, so callers can index [0]. The employee's original bytes are not
    touched; these are copies.

    The two limits are deliberately separate. Short-circuiting the validation
    loop once ``render_limit`` pages exist is the obvious optimization and
    silently downgrades the preflight: an 11-page PDF or an oversized page 7
    would then pass at upload time and only fail later, during generation,
    after the whole report had been typed in.

    Suffix-driven dispatch means a mislabelled file (a PDF named .jpg) takes
    the image branch and fails as an unreadable image. That is the visible
    outcome, not a silent one.
    """
    suffix = Path(filename).suffix.casefold()
    if suffix == ".pdf":
        import fitz

        pages: list[bytes] = []
        try:
            with fitz.open(stream=file_bytes, filetype="pdf") as document:
                if document.page_count > max_pages:
                    raise ExpenseReportError(
                        f"{filename} has {document.page_count} pages; the limit is {max_pages}."
                    )
                for page_number, page in enumerate(document, 1):
                    # Predict the raster size from the page's declared POINTS
                    # (72 per inch) at the 150 DPI used below, and check it
                    # BEFORE get_pixmap allocates anything. A 300 KB PDF can
                    # declare a 200-inch page; validating the pixmap after the
                    # fact validates it after the memory is already gone.
                    # ceil() so a fractional page never rounds under the limit.
                    width = ceil(page.rect.width * 150 / 72)
                    height = ceil(page.rect.height * 150 / 72)
                    _validate_receipt_dimensions(
                        width,
                        height,
                        label=f"{filename} page {page_number}",
                    )
                    # `continue`, not `break`: the remaining pages still have
                    # to be size-checked even though they will not be
                    # rendered. See the docstring.
                    if render_limit is not None and len(pages) >= render_limit:
                        continue
                    pixmap = page.get_pixmap(dpi=150, alpha=False)
                    image = Image.open(BytesIO(pixmap.tobytes("png")))
                    pages.append(_compact_receipt_image(image))
        # Re-raise our OWN errors untouched, then blanket-wrap the rest. The
        # order is the whole point: without the first clause the specific
        # "11 pages; the limit is 10" message raised inside the loop would be
        # swallowed and replaced by the generic "not a readable PDF", telling
        # the employee to re-scan a file that scanned perfectly well.
        except ExpenseReportError:
            raise
        except Exception as exc:
            raise ExpenseReportError(f"{filename} is not a readable PDF receipt.") from exc
        if not pages:
            raise ExpenseReportError(f"{filename} does not contain a receipt page.")
        return pages

    # Registered lazily and only for HEIC suffixes: pillow_heif is a heavy
    # import and every non-Apple receipt would pay for it at module load.
    # thumbnails=False so Pillow decodes the FULL image -- an iPhone HEIC
    # embeds a small preview, and opening that instead yields a receipt too
    # low-resolution to read, with no error to say why.
    if suffix in {".heic", ".heif", ".hif"}:
        from pillow_heif import register_heif_opener

        register_heif_opener(thumbnails=False)
    try:
        pages = []
        with Image.open(BytesIO(file_bytes)) as image:
            # Animated GIF and multi-page TIFF receipts are real. getattr with
            # a default because single-frame formats do not define n_frames at
            # all -- a bare image.n_frames would raise AttributeError on every
            # ordinary JPEG and be reported as an unreadable receipt.
            frame_count = int(getattr(image, "n_frames", 1))
            if frame_count > max_pages:
                raise ExpenseReportError(
                    f"{filename} has {frame_count} frames; the limit is {max_pages}."
                )
            for frame_index in range(frame_count):
                image.seek(frame_index)
                _validate_receipt_dimensions(
                    image.width,
                    image.height,
                    label=filename,
                )
                if render_limit is None or len(pages) < render_limit:
                    pages.append(_compact_receipt_image(image.copy()))
        if not pages:
            raise ExpenseReportError(f"{filename} does not contain a receipt image.")
        return pages
    except ExpenseReportError:
        raise
    except Exception as exc:
        raise ExpenseReportError(f"{filename} is not a readable receipt image.") from exc


def _compact_receipt_image(source: Image.Image) -> bytes:
    """Normalize one decoded frame into a compact, print-readable JPEG.

    Everything here is required by the destination, not cosmetic: the result is
    embedded in the RECEIPTS worksheet, rendered into the PDF the approver
    reads, and mailed. JPEG has no alpha channel and no orientation tag, so
    transparency and EXIF rotation must both be resolved here or they resolve
    themselves badly.
    """
    # Re-validated even though every caller already checked. Not redundant for
    # PDFs: the caller checks an ESTIMATE derived from the page's declared
    # points, this checks the raster that actually came back.
    _validate_receipt_dimensions(source.width, source.height)
    # Apply the EXIF orientation tag now. JPEG output keeps no such tag, so a
    # sideways phone photo left untransposed is a sideways receipt in the
    # approver's PDF -- rotated, but never reported as anything.
    image = ImageOps.exif_transpose(source)
    # Composite onto WHITE before dropping alpha. A bare .convert("RGB")
    # discards the alpha channel and keeps whatever RGB the transparent pixels
    # carried, which for a scanned or screenshotted receipt is usually black --
    # the page background becomes a black field and the text disappears, with
    # no error anywhere. app/ocr.py carries the same fix for the same reason.
    #
    # Convert to RGBA FIRST, then use its alpha as the mask. That is what makes
    # this work for a palette receipt, and it is why the mode test and the mask
    # must not be split apart again.
    #
    # The previous version tested `image.mode in {"RGBA", "LA"} or
    # "transparency" in image.info` and then took the mask from
    # `image.getchannel("A") if "A" in image.getbands() else None`. A PNG-8 or
    # GIF receipt is mode "P": it PASSED the test via info["transparency"], but
    # its bands are ("P",) with no "A", so the mask was None and
    # paste(..., mask=None) is a plain overwrite. The flatten silently did not
    # happen. Measured on a transparent-background palette receipt: 100% of the
    # output was near-black -- the text destroyed along with the background --
    # and it went into the approver's submission PDF with no error anywhere.
    # Mode "PA" missed the old test entirely.
    #
    # Keep this identical to the flatten in app/ocr.py. The two are the same fix
    # for the same reason, and they had already drifted apart once.
    if image.mode in {"RGBA", "LA", "PA"} or (
        image.mode == "P" and "transparency" in image.info
    ):
        background = Image.new("RGB", image.size, "white")
        rgba = image.convert("RGBA")
        background.paste(rgba, mask=rgba.split()[-1])
        image = background
    else:
        image = image.convert("RGB")
    # contain(), so an already-small receipt is never UPSCALED into a blurry
    # larger file. tests/test_expense_report.py pins that a 480x720 source
    # comes back 480x720.
    if (
        image.width > _ATTACHMENT_MAX_SIZE[0]
        or image.height > _ATTACHMENT_MAX_SIZE[1]
    ):
        image = ImageOps.contain(image, _ATTACHMENT_MAX_SIZE)
    buffer = BytesIO()
    image.save(buffer, format="JPEG", quality=78, optimize=True, progressive=True)
    return buffer.getvalue()


def _validate_receipt_dimensions(
    width: int,
    height: int,
    *,
    label: str = "Receipt image",
) -> None:
    """Reject invalid or decompression-heavy pages before raster allocation.

    Call this on DECLARED dimensions, before any decode. Pillow's own
    decompression-bomb guard only trips near 178 MP, so a frame in the
    40-178 MP band materializes hundreds of MB on a shared container before
    anything objects -- once per frame. ``label`` names the page so a ten-page
    PDF says which one is the problem.
    """
    if width <= 0 or height <= 0:
        raise ExpenseReportError(f"{label} has invalid dimensions.")
    if width * height > _MAX_RECEIPT_PIXELS:
        raise ExpenseReportError(
            f"{label} is too large ({width}×{height}); use a smaller image or PDF page."
        )


def _verify_template_anchors(sheet) -> None:
    """Fail closed if the supplied template no longer matches our coordinates.

    Every row number in this module is a hardcoded coordinate into a workbook
    the customer supplies. If a future revision of that form inserts a row, the
    fills still succeed and produce a document that looks finished with each
    section's data one row out of place -- amounts under the wrong heading,
    coding in the wrong column. There is no exception to catch for that.

    These four cells are chosen to straddle the ranges: B2 above the header
    block, B20 and B41 immediately above the two expense blocks, J60 below the
    totals. A shift anywhere between them moves at least one of these labels.
    The message says "no data was written" because none was -- this runs before
    the first fill.

    Substring matching, not equality, because the template's own B2 carries a
    trailing revision note ("... - 2025 (Effective October 2025)"). Tightening
    this to equality would reject the very template it is meant to accept, on
    the next revision of that note.
    """
    anchors = {
        "B2": "REIMBURSEMENT OF EXPENSES",
        "B20": "MISCELLANEOUS",
        "B41": "ENTERTAINMENT",
        "J60": "TOTAL REIMBURSEMENT",
    }
    for cell, expected in anchors.items():
        if expected not in str(sheet[cell].value or ""):
            raise ExpenseReportError(
                f"The reimbursement template changed near cell {cell}; no data was written."
            )


def _fill_report_header(
    sheet,
    details: ExpenseReportDetails,
    mileage_items: Sequence[MileageItem],
    image_buffers: list[BytesIO],
) -> None:
    """Fill the header block, the mileage-rate scaffolding, and the signature.

    MUST run before the mileage rows are filled: the loop below pre-fills all
    eight rows with zeros and a rate formula, so running it afterwards would
    erase the real figures. build_expense_workbook documents the ordering.

    ``image_buffers`` is an out-parameter that keeps the signature's stream
    referenced until the workbook is saved; nothing reads it.
    """
    # Refreshed from the report year rather than left at the template's
    # original October 2025 text, so an old year is not printed on a current
    # submission. This overwrites an anchor _verify_template_anchors just
    # checked, which is safe because verification already ran.
    sheet["B2"] = f"REIMBURSEMENT OF EXPENSES - 01.01.{details.report_date:%Y}"
    _write_excel_text(sheet["C5"], details.employee_name)
    _write_excel_text(sheet["G5"], details.employee_number)
    _write_excel_text(sheet["K5"], details.employee_home_bu)
    sheet["P5"] = details.report_date
    sheet["P5"].number_format = "m/d/yyyy"

    # The printed heading must not claim a single rate when the rows use two.
    # A report spanning the 2026-07-01 IRS change legitimately mixes $0.725 and
    # $0.76 rows, and a heading naming one of them would read as an arithmetic
    # error on the approver's copy. The set of DISTINCT rates decides:
    # one -> name it, several -> say "APPLICABLE", none -> fall back to the
    # report date, and if even that is unconfigured, say nothing specific
    # rather than print a rate no row used.
    rates = {
        rate
        for item in mileage_items
        if (rate := irs_business_mileage_rate(item.transaction_date)) is not None
    }
    default_rate = irs_business_mileage_rate(details.report_date)
    if len(rates) == 1:
        displayed_rate = next(iter(rates))
        sheet["B6"] = f"GAS MILEAGE @ ${displayed_rate} PER MILE"
    elif len(rates) > 1:
        sheet["B6"] = "GAS MILEAGE @ APPLICABLE IRS RATE"
    elif default_rate is not None:
        sheet["B6"] = f"GAS MILEAGE @ ${default_rate} PER MILE"
    else:
        sheet["B6"] = "GAS MILEAGE @ IRS BUSINESS RATE"
    # Scaffold ALL eight mileage rows with 0 miles and a live rate formula.
    # Used rows are overwritten immediately afterwards with their own
    # travel-date rate; unused rows stay 0 x rate = $0.00 so H18's SUM covers
    # the whole range without picking up whatever the template left behind.
    #
    # `default_rate or Decimal("0")` is only ever seen by rows that stay empty:
    # a real mileage row with an unconfigured rate is blocked by validation, so
    # a 0 rate can never silently zero out a claimed trip.
    formula_rate = default_rate or Decimal("0")
    for row in MILEAGE_ROWS:
        sheet.cell(row=row, column=7, value=0)
        sheet.cell(
            row=row,
            column=8,
            value=f"=ROUND(G{row}*{formula_rate},2)",
        ).number_format = '"$"#,##0.00'

    # Both boxes are written on every run, one of them blank. Writing only the
    # chosen one would leave a stale "x" from the template or from a previous
    # revision of the form, i.e. a check routed to two destinations at once.
    sheet["B62"] = "x" if details.mail_destination == "home" else ""
    sheet["B64"] = "x" if details.mail_destination == "satellite" else ""
    if details.mail_destination == "satellite":
        _write_excel_text(sheet["F64"], details.satellite_office)

    signature_buffer = BytesIO(employee_signature_png(details.employee_name))
    image_buffers.append(signature_buffer)
    signature_image = ExcelImage(signature_buffer)
    # Display size in the signature line's cell box, independent of the
    # rendered pixel size. employee_signature_png guarantees the image is
    # complete (never clipped) at whatever resolution it chose, so scaling it
    # down to this box only changes how large it prints.
    signature_image.width = 170
    signature_image.height = 30
    sheet.add_image(signature_image, "C66")
    # Two-digit year, matching the form's own printed convention. %-d/%-m are
    # not portable, hence the explicit .month/.day, and `% 100:02d` rather than
    # a %y format so the source of the two digits is visible.
    sheet["D66"] = (
        f"Date: {details.report_date.month}/{details.report_date.day}/"
        f"{details.report_date.year % 100:02d}"
    )
    sheet["D66"].font = Font(name="Arial", size=8)
    sheet["D66"].alignment = Alignment(shrinkToFit=True)
    # Blanked, not skipped: E66 sits inside the signature/date line and must
    # not carry template residue next to a generated signature.
    sheet["E66"] = ""
    # The APPROVER's signature and printed-name cells are deliberately left
    # untouched. This tool prepares a draft for a human to approve; filling
    # them would fabricate an approval.
    _write_excel_text(sheet["C68"], details.employee_name)
    sheet["C68"].font = Font(name="Arial", size=12)
    sheet["C68"].alignment = Alignment(shrinkToFit=True)


def employee_signature_png(employee_name: str) -> bytes:
    """Render a deterministic cursive employee-name signature preview.

    The font is shrunk to fit the canvas rather than left at a fixed size and
    clipped. This is a correctness property, not an aesthetic one: the rendered
    image is embedded at C66 of the official JDE reimbursement form that goes to
    the approver, and the employee attests to it via "I confirm this generated
    signature represents me". A fixed 112pt size drew any long name straight
    past the 1600px canvas edge, and the subsequent crop clamp silently returned
    a cut-off signature — so the employee could attest to a truncated version of
    their own name on a financial document.
    """
    name = " ".join(str(employee_name or "").split())
    if not name or len(name) > _SIGNATURE_MAX_NAME_CHARS:
        raise ExpenseReportError("Enter a valid employee name for the signature.")
    font_path = next(
        (candidate for candidate in _SIGNATURE_FONT_CANDIDATES if candidate.is_file()),
        None,
    )
    if font_path is None:
        raise ExpenseReportError(
            "The cursive signature font is unavailable in this deployment."
        )

    canvas = Image.new("RGBA", _SIGNATURE_CANVAS, (255, 255, 255, 0))
    drawing = ImageDraw.Draw(canvas)
    max_width = _SIGNATURE_CANVAS[0] - _SIGNATURE_MARGIN[0] * 2
    max_height = _SIGNATURE_CANVAS[1] - _SIGNATURE_MARGIN[1] * 2

    size = _SIGNATURE_MAX_FONT_SIZE
    while True:
        font = ImageFont.truetype(str(font_path), size)
        box = drawing.textbbox((0, 0), name, font=font)
        width = max(1, box[2] - box[0])
        height = max(1, box[3] - box[1])
        if size <= _SIGNATURE_MIN_FONT_SIZE or (
            width <= max_width and height <= max_height
        ):
            break
        # Step toward a proportional fit, but always shrink by at least 1pt so a
        # pathological metric can never stall this loop.
        scale = min(max_width / width, max_height / height)
        size = max(_SIGNATURE_MIN_FONT_SIZE, min(size - 1, int(size * scale)))

    if width > max_width or height > max_height:
        # Reached the legibility floor and it still does not fit. Refuse rather
        # than emit a clipped signature: a visible error the employee can act on
        # is always better than a financial document carrying a silently
        # truncated attestation of their name.
        raise ExpenseReportError(
            "That name is too long to render as a signature. Use a shorter form "
            "of the name (for example, a middle initial instead of full middle "
            "names)."
        )

    drawing.text(
        (_SIGNATURE_MARGIN[0] - box[0], _SIGNATURE_MARGIN[1] - box[1]),
        name,
        font=font,
        fill=(0, 0, 0, 255),
    )
    cropped = canvas.crop(
        (
            0,
            0,
            min(_SIGNATURE_CANVAS[0] - 1, width + _SIGNATURE_MARGIN[0] * 2),
            min(_SIGNATURE_CANVAS[1] - 1, height + _SIGNATURE_MARGIN[1] * 2),
        )
    )
    if cropped.width > 1100:
        cropped = ImageOps.contain(cropped, (1100, 190))
    payload = BytesIO()
    cropped.save(payload, format="PNG")
    return payload.getvalue()


def _mileage_grouped_for_form(
    items: Sequence[MileageItem],
) -> list[MileageItem]:
    """Group mileage rows by coding, then date, for an auditable form.

    The template's own note directs grouping by job number and code. entry_id
    is the final tiebreaker so the order is TOTAL and deterministic: two
    identical rows must not swap places between generations, or the content
    fingerprint would change without the operator changing anything.

    ``date.max`` for a missing date parks undated rows last instead of raising
    on a None comparison. Validation blocks them long before this runs; the
    fallback exists so a direct call cannot crash mid-sort.
    """
    return sorted(
        items,
        key=lambda item: (
            item.allocation.job_number,
            item.allocation.account_cost_type,
            item.allocation.cost_code_or_wo_type,
            item.transaction_date or date.max,
            item.entry_id,
        ),
    )


def _fill_mileage_row(sheet, row: int, item: MileageItem) -> None:
    """Write one validated mileage row. Assumes validation already passed."""
    miles = parse_mileage(item.miles)
    rate = irs_business_mileage_rate(item.transaction_date)
    # A restatement of what validate_expense_report already guaranteed, not a
    # gate. It exists so a future direct caller that skips validation fails
    # here, loudly, instead of writing float(None) or a formula with the word
    # "None" in it into a financial form.
    assert item.transaction_date is not None and miles is not None and rate is not None
    sheet.cell(row=row, column=2, value=item.transaction_date).number_format = "m/d/yyyy"
    _write_excel_text(sheet.cell(row=row, column=3), item.purpose)
    _write_excel_text(sheet.cell(row=row, column=6), item.destination)
    sheet.cell(row=row, column=7, value=float(miles)).number_format = "0.00"
    # THIS row's travel-date rate, overwriting the report-date scaffold from
    # _fill_report_header. A live formula rather than a computed number so the
    # workbook stays editable: correcting the miles in column G updates the
    # reimbursement instead of silently contradicting it.
    sheet.cell(
        row=row,
        column=8,
        value=f"=ROUND(G{row}*{rate},2)",
    ).number_format = '"$"#,##0.00'
    _fill_allocation(sheet, row, item.allocation)


def _items_grouped_for_form(items: list[ExpenseItem]) -> list[ExpenseItem]:
    """Group rows by section and coding, preserving upload order within a group.

    sorted() is STABLE, which is what preserves upload order inside a coding
    group -- replacing it with anything that reorders equal keys changes the
    printed form and the receipt order for no reason, and moves the content
    fingerprint with it.

    The key spans the COMPLETE allocation, not just the three written columns.
    Two lines that differ only in an unwritten legacy field are genuinely
    different coding strings, and merging them into one visual group would
    misrepresent the report to whoever audits it.
    """
    section_order = {
        EXPENSE_SECTION_MISC: 0,
        EXPENSE_SECTION_ENTERTAINMENT: 1,
    }

    def key(item: ExpenseItem) -> tuple[object, ...]:
        allocation = item.allocation
        return (
            # 9 for an unrecognized section: it sorts last and, crucially,
            # keeps sorted() from raising on a value the caller has not
            # validated yet. It never reaches the form -- neither the
            # Miscellaneous nor the Entertainment slice below matches it, and
            # validation blocks it long before this point.
            section_order.get(item.section, 9),
            allocation.kind,
            allocation.job_number,
            allocation.service_center,
            allocation.account_cost_type,
            allocation.cost_code_or_wo_type,
            allocation.work_order_number,
            allocation.company_number,
            allocation.department_number,
            allocation.ou_number,
            allocation.gl_account_number,
        )

    return sorted(items, key=key)


def _fill_expense_row(sheet, row: int, item: ExpenseItem, *, entertainment: bool) -> None:
    """Write one validated reimbursement line into an expense row.

    ``entertainment`` selects the column-F behaviour, not the row range -- the
    caller has already chosen which block of rows to write into. Column F is
    the entertainment CONTACT and is left untouched for Miscellaneous rows,
    where the same column means something else on the official form.
    """
    amount = parse_expense_amount(item.amount)
    # See _fill_mileage_row: a restatement of validation, so a caller that
    # skipped it cannot write float(None) into a currency cell.
    assert item.transaction_date is not None and amount is not None
    sheet.cell(row=row, column=2, value=item.transaction_date).number_format = "m/d/yyyy"
    _write_excel_text(sheet.cell(row=row, column=3), item.description)
    if entertainment:
        _write_excel_text(sheet.cell(row=row, column=6), item.contact_name)
    # Written as a NUMBER, in deliberate contrast to every text cell around it.
    # The =SUM(H24:H38) written by build_expense_workbook totals this column,
    # and a currency-formatted STRING would display identically while
    # contributing zero to that total -- a form whose rows and total disagree,
    # with no error. Do not route this through _write_excel_text.
    # parse_expense_amount already quantized to two places, so the float
    # conversion is exact at reimbursement magnitudes.
    sheet.cell(row=row, column=8, value=float(amount)).number_format = '"$"#,##0.00'
    _fill_allocation(sheet, row, item.allocation)


def _fill_allocation(sheet, row: int, allocation: ExpenseAllocation) -> None:
    """Write columns I/J/K for one row, or refuse the whole workbook.

    The SECOND enforcement of job-only coding; allocation_problems is the
    first. Duplicated on purpose: this is the last point before a value reaches
    the sheet, and it fails the entire generation rather than skipping the row,
    because a skipped row is an uncoded line on a submitted form.
    """
    if allocation.kind != ALLOCATION_JOB:
        raise ExpenseReportError(
            "Expense reports use job coding only; Work Order and Other Expenses "
            "columns must remain blank."
        )
    # The catalog option is a DESCRIPTIVE label ("RRH-695400022-O&M"); column I
    # takes only the numeric 695.../VI... identifier inside it. Writing the
    # whole label would put a human-readable string where JDE expects a job
    # number, and it would look deliberate on the printed form.
    job_identifier = job_number_identifier(allocation.job_number)
    if not job_identifier:
        raise ExpenseReportError("The selected job number has no usable identifier.")
    _write_code(sheet, row, 9, job_identifier)
    _write_code(sheet, row, 10, allocation.account_cost_type)
    _write_code(sheet, row, 11, allocation.cost_code_or_wo_type)


def _write_code(sheet, row: int, column: int, value: str) -> None:
    """Write an accounting code as Excel TEXT, preserving leading zeros.

    Two separate mechanisms, both needed. _write_excel_text stores the value as
    a string, which protects the GENERATED file. The "@" number format protects
    the file after the employee opens it: the workbook is deliberately
    editable, and in a general-format cell Excel converts a retyped numeric
    code like "05490" to 5490 on commit -- the leading zero disappears from the
    submitted form with nothing on screen saying it did.
    """
    cell = sheet.cell(row=row, column=column)
    _write_excel_text(cell, value)
    cell.number_format = "@"


def _write_excel_text(cell, value: object) -> None:
    """Write an explicit string so user text can never become an Excel formula.

    Used for EVERY user-editable text cell. A description, merchant or employee
    name beginning with "=", "+", "-" or "@" is otherwise stored as a formula:
    the receipt line vanishes behind a computed value or an error code, and the
    generated document silently stops saying what the employee typed.
    tests/test_expense_report.py drives each of those four prefixes through the
    whole workbook.

    Setting data_type after value is required, not decorative -- assigning
    .value re-infers the type, so the two lines cannot be swapped or merged.
    """
    cell.value = str(value or "").strip()
    cell.data_type = "s"


def _build_receipt_sheet(sheet, items: list[ExpenseItem], buffers: list[BytesIO]) -> None:
    """Lay out one printed page per receipt page, in the form's own row order.

    ``items`` must already be in _items_grouped_for_form order; the numbering
    here follows first appearance, so the packet reads in the same sequence as
    the rows above it. ``buffers`` is an out-parameter holding every image
    stream alive until the workbook is saved.

    THE PAGE GEOMETRY IS ARITHMETIC, not taste. Letter portrait is 612 x 792
    points. With the 0.35" side margins set below, the printable width is
    612 - 2*0.35*72 = 561.6pt, and twelve columns at Excel width 8.2
    (8.2*7 + 5 = 62.4px = 46.8pt each) come to exactly 561.6pt. That is why the
    column count, the column width and the side margins are one decision:
    change any of them alone and the receipt block no longer fills, or no
    longer fits, the page it was measured against.

    Each receipt page occupies _RECEIPT_PAGE_ROWS rows with an explicit page
    break on the boundary, which is what keeps one receipt per printed page.
    tests/test_expense_report.py renders the real combined PDF and asserts the
    page count and which receipt lands on which page, so a change here is
    caught -- but only where LibreOffice is installed.
    """
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    # fitToHeight = 0 means "as many pages tall as needed", NOT "zero pages".
    # Setting it to 1 would squeeze an entire multi-receipt packet onto one
    # sheet, shrinking every receipt to illegibility -- and it would still
    # render without complaint.
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    # Both halves are required: page_setup carries the numbers, and this flag
    # is what makes Excel/LibreOffice honour them instead of the default
    # percentage scaling. Setting only the numbers fails silently.
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.35
    sheet.page_margins.right = 0.35
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4
    # chr(64 + n) spans A..L for n in 1..12. Correct only while the sheet is
    # twelve columns wide -- past column 26 this produces punctuation, not
    # column letters, and openpyxl would happily create a dimension entry for
    # a column that does not exist.
    for column in range(1, 13):
        sheet.column_dimensions[chr(64 + column)].width = 8.2

    # One block per SOURCE receipt, not per reimbursement line. This is what
    # attaches a split receipt's image exactly once while the form above shows
    # each of its lines separately.
    receipt_groups = _receipt_groups(items)
    # Counts printed PAGES, unlike receipt_number which counts receipts. A
    # multi-page PDF receipt advances it several times, which is what keeps the
    # row offsets and the page breaks aligned when receipt counts and page
    # counts differ.
    block_index = 0
    for receipt_number, group in enumerate(receipt_groups, 1):
        # group[0] supplies the shared receipt identity -- date, merchant,
        # filename, bytes. Every line in a group is validated to carry the same
        # uploaded file, so any member would do.
        item = group[0]
        # No render_limit here: the packet needs every page, whereas the upload
        # preflight only needed page one.
        pages = receipt_attachment_pages(item.file_bytes, item.filename)
        for page_number, image_bytes in enumerate(pages, 1):
            start = block_index * _RECEIPT_PAGE_ROWS + 1
            end = start + _RECEIPT_PAGE_ROWS - 1
            sheet.merge_cells(start_row=start, start_column=1, end_row=start, end_column=12)
            sheet.cell(start, 1).value = (
                f"Receipt {receipt_number} of {len(receipt_groups)}"
                + (
                    f" · {len(group)} reimbursement lines"
                    if len(group) > 1
                    else ""
                )
                + (f" · page {page_number} of {len(pages)}" if len(pages) > 1 else "")
            )
            header_font = copy(sheet.cell(start, 1).font)
            header_font.bold = True
            header_font.size = 14
            header_font.color = "092B24"
            sheet.cell(start, 1).font = header_font
            sheet.merge_cells(
                start_row=start + 1,
                start_column=1,
                end_row=start + 1,
                end_column=12,
            )
            # The REIMBURSED total across this receipt's lines, which is
            # routinely LESS than the receipt's printed total -- the approved
            # reference report claims one business line out of a larger
            # purchase. Labelling it "reimbursable" is deliberate: an auditor
            # comparing the header to the image must see that the difference is
            # intended, not a transcription error.
            amount = sum(
                (parse_expense_amount(line.amount) or Decimal("0") for line in group),
                Decimal("0"),
            )
            detail = " · ".join(
                part
                for part in (
                    item.transaction_date.strftime("%m/%d/%Y") if item.transaction_date else "",
                    item.merchant_name.strip(),
                    f"${amount:,.2f} reimbursable",
                    Path(item.filename).name,
                )
                if part
            )
            sheet.cell(start + 1, 1).value = detail
            detail_font = copy(sheet.cell(start + 1, 1).font)
            detail_font.size = 9
            detail_font.color = "557F7F"
            sheet.cell(start + 1, 1).font = detail_font

            buffer = BytesIO(image_bytes)
            buffers.append(buffer)
            excel_image = ExcelImage(buffer)
            # Pixel box inside one _RECEIPT_PAGE_ROWS block, leaving room for
            # the two header rows above the anchor and a bottom margin.
            available_width = 690
            available_height = 760
            # The trailing 1 caps the scale at 1.0 so a small receipt is
            # centred at its own size rather than blown up. Dropping it would
            # upscale a 400px phone screenshot to 690px of blur on the page the
            # approver actually reads.
            scale = min(
                available_width / excel_image.width,
                available_height / excel_image.height,
                1,
            )
            excel_image.width = int(excel_image.width * scale)
            excel_image.height = int(excel_image.height * scale)
            sheet.add_image(excel_image, f"A{start + 3}")
            # `if block_index` skips the break before the FIRST block. A break
            # at row 0 would either be ignored or emit a blank leading page,
            # and a blank first page in the approver's packet reads as a
            # missing form.
            if block_index:
                sheet.row_breaks.append(Break(id=start - 1))
            for row in range(start, end + 1):
                sheet.row_dimensions[row].height = 14
            sheet.row_dimensions[start].height = 24
            sheet.row_dimensions[start + 1].height = 18
            block_index += 1

    # An explicit print area, so a stray cell outside the blocks can never add
    # a page to the packet. max(1, ...) keeps the range syntactically valid if
    # no block was written at all -- "A1:L0" is not a range and would surface
    # as a corrupt workbook rather than an empty sheet.
    last_row = max(1, block_index * _RECEIPT_PAGE_ROWS)
    sheet.print_area = f"A1:L{last_row}"
    sheet.oddFooter.center.text = "Receipt attachments"
    sheet.oddFooter.right.text = "Page &P of &N"


def _receipt_source_id(item: ExpenseItem) -> str:
    """Return the uploaded-receipt identity shared by split lines.

    THE definition of "same uploaded receipt" for the whole module -- grouping,
    attachment de-duplication, receipt counting and the duplicate warning all
    agree because they all ask this. Do not re-derive it at a call site.

    Falls back to receipt_id for an unsplit line, which has no
    source_receipt_id. The final "" fallback is unreachable in practice
    because validate_expense_report rejects an empty receipt_id; if it were
    reachable, every such line would group under one key and share a single
    attachment, which is why that validation is not optional.
    """
    return str(item.source_receipt_id or item.receipt_id or "").strip()


def _receipt_groups(items: Sequence[ExpenseItem]) -> list[list[ExpenseItem]]:
    """Group reimbursement lines by source receipt, preserving first use order.

    Relies on dict insertion order, so a group appears at the position of its
    FIRST line. Feed it the form-ordered list and the receipt pages come out in
    the same sequence as the rows; feed it the raw list and the packet stops
    matching the form.
    """
    grouped: dict[str, list[ExpenseItem]] = {}
    for item in items:
        grouped.setdefault(_receipt_source_id(item), []).append(item)
    return list(grouped.values())


def _unique_receipt_count(items: Sequence[ExpenseItem]) -> int:
    """Count uploaded receipts, not reimbursement lines.

    This is the number quoted as "Receipts" in the approval email, so it must
    match what the approver can count in the attached packet. len(items) would
    over-report every split receipt.
    """
    return len({_receipt_source_id(item) for item in items})


def _looks_like_email(value: str) -> bool:
    """Cheap shape check for the approver address, not RFC 5322 validation.

    Deliberately permissive: it catches a missing "@" or a stray space, which
    are what a person actually mistypes, and refuses to reject an unusual but
    real corporate address. A stricter pattern that rejects a valid approver
    blocks a finished report with no way around it.

    The [^\\s@] classes exclude CR and LF, which is also what keeps a header
    injection out of the .eml the address is later used to build.
    """
    text = str(value or "").strip()
    return bool(
        re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", text)
        and len(text) <= 254
    )


def _deduplicate(values: list[str]) -> list[str]:
    """De-duplicate while preserving first-seen order, dropping empties.

    Order preservation is the point: validation messages are emitted in the
    order the operator meets the fields, and set() would scramble them into a
    different order on every run.
    """
    return list(dict.fromkeys(value for value in values if value))
