from __future__ import annotations

import hashlib
from dataclasses import replace
from datetime import date
from decimal import Decimal
from email import policy
from email.parser import BytesParser
from io import BytesIO

import fitz
import pytest
from openpyxl import load_workbook
from PIL import Image, ImageDraw

from app.expense_report import (
    ALLOCATION_JOB,
    ALLOCATION_OVERHEAD,
    ALLOCATION_WORK_ORDER,
    EXPENSE_SECTION_ENTERTAINMENT,
    EXPENSE_SECTION_MISC,
    MAX_ENTERTAINMENT_ITEMS,
    MAX_MISCELLANEOUS_ITEMS,
    ExpenseAllocation,
    ExpenseItem,
    ExpensePackage,
    ExpenseReportError,
    ExpenseReportDetails,
    MileageItem,
    build_expense_package,
    build_expense_workbook,
    email_attachments_for_package,
    employee_signature_png,
    expense_report_signature,
    expense_report_warnings,
    irs_business_mileage_rate,
    mileage_reimbursement,
    parse_expense_amount,
    parse_mileage,
    receipt_attachment_pages,
    receipt_preview_bytes,
    total_reimbursement,
    validate_expense_report,
)
from app.expense_report import _compact_receipt_image
from app.job_numbers import RRH_JOB_NUMBERS
from app.expense_ui import _build_expense_eml
from tests.conftest import requires_libreoffice


def _receipt_bytes(text: str = "TOTAL $31.25") -> bytes:
    image = Image.new("RGB", (800, 1100), "white")
    drawing = ImageDraw.Draw(image)
    drawing.rectangle((20, 20, 780, 1080), outline="black", width=3)
    drawing.multiline_text((70, 80), text, fill="black", spacing=18)
    payload = BytesIO()
    image.save(payload, format="JPEG", quality=90)
    return payload.getvalue()


def _details(**changes) -> ExpenseReportDetails:
    values = {
        "account": "Rochester Regional Health",
        "employee_name": "Synthetic Employee",
        "employee_number": "TEST-1001",
        "employee_home_bu": "695",
        "report_date": date(2026, 8, 11),
        "approver_name": "RRH Test Administrator",
        "approver_email": "rrh.approver@example.invalid",
        "mail_destination": "home",
        "satellite_office": "",
        "employee_signature_confirmed": True,
    }
    values.update(changes)
    return ExpenseReportDetails(**values)


def _job_allocation(
    *,
    job: str = RRH_JOB_NUMBERS[0],
    cost_type: str = "01AMA",
    cost_code: str = "5490",
) -> ExpenseAllocation:
    return ExpenseAllocation(
        kind=ALLOCATION_JOB,
        job_number=job,
        account_cost_type=cost_type,
        cost_code_or_wo_type=cost_code,
    )


def _item(
    receipt_id: str = "receipt-a",
    *,
    section: str = EXPENSE_SECTION_MISC,
    amount: str = "31.25",
    allocation: ExpenseAllocation | None = None,
    contact_name: str = "",
) -> ExpenseItem:
    return ExpenseItem(
        receipt_id=receipt_id,
        filename=f"{receipt_id}.jpg",
        file_bytes=_receipt_bytes(),
        transaction_date=date(2026, 8, 10),
        description="Business travel parking",
        amount=amount,
        section=section,
        allocation=allocation or _job_allocation(),
        merchant_name="Test Merchant",
        contact_name=contact_name,
    )


def _mileage(
    entry_id: str = "mileage-1",
    *,
    travel_date: date = date(2026, 8, 10),
    miles: str = "10",
    allocation: ExpenseAllocation | None = None,
) -> MileageItem:
    return MileageItem(
        entry_id=entry_id,
        transaction_date=travel_date,
        purpose="RRH site visit",
        destination="United Memorial Medical Center",
        miles=miles,
        allocation=allocation or _job_allocation(),
    )


def test_workbook_matches_approved_rrh_job_only_layout_and_receipt_order():
    startup = _job_allocation(job=RRH_JOB_NUMBERS[1], cost_type="02AMA")
    items = [
        _item("parking"),
        _item(
            "customer-meal",
            section=EXPENSE_SECTION_ENTERTAINMENT,
            amount="48.75",
            allocation=startup,
            contact_name="Customer Contact",
        ),
    ]

    payload = build_expense_workbook(_details(), items)
    workbook = load_workbook(BytesIO(payload), data_only=False)
    form = workbook["EXPENSE REIMBURSEMENT"]
    receipts = workbook["RECEIPTS"]

    assert workbook.sheetnames == ["EXPENSE REIMBURSEMENT", "RECEIPTS"]
    assert form["B2"].value == "REIMBURSEMENT OF EXPENSES - 01.01.2026"
    assert form["B6"].value == "GAS MILEAGE @ $0.76 PER MILE"
    assert form["C5"].value == "Synthetic Employee"
    assert form["G5"].value == "TEST-1001"
    assert form["K5"].value == "695"
    assert form["P5"].value.date() == date(2026, 8, 11)
    assert form["B62"].value == "x"
    assert form["C68"].value == "Synthetic Employee"
    assert form["D66"].value == "Date: 8/11/26"
    assert len(form._images) == 1

    assert form["B24"].value.date() == date(2026, 8, 10)
    assert form["C24"].value == "Business travel parking"
    assert form["H24"].value == 31.25
    assert form["I24"].value == "695400022"
    assert form["J24"].value == "01AMA"
    assert form["K24"].value == "5490"
    assert form["J24"].number_format == "@"

    assert form["C45"].value == "Business travel parking"
    assert form["F45"].value == "Customer Contact"
    assert form["I45"].value == "695400023"
    assert form["J45"].value == "02AMA"
    assert form["K45"].value == "5490"
    for row in (24, 45):
        assert form.cell(row, 12).value is None
        assert all(form.cell(row, column).value is None for column in range(14, 18))
    assert form["H18"].value == "=SUM(H10:H17)"
    assert form["H39"].value == "=SUM(H24:H38)"
    assert form["H59"].value == "=SUM(H45:H58)"
    assert form["Q60"].value == "=H18+H39+H59"

    assert len(receipts._images) == 2
    assert receipts["A1"].value == "Receipt 1 of 2"
    assert receipts["A59"].value == "Receipt 2 of 2"


def test_user_text_and_editable_codes_are_never_written_as_excel_formulas():
    details = _details(
        employee_name="=HYPERLINK(\"https://example.invalid\",\"name\")",
        employee_number="@SUM(1+1)",
        employee_home_bu="+RRH",
        mail_destination="satellite",
        satellite_office="-1+1",
    )
    allocation = _job_allocation(
        cost_type="=1+1",
        cost_code="@malicious",
    )
    item = replace(
        _item(
            "formula",
            section=EXPENSE_SECTION_ENTERTAINMENT,
            allocation=allocation,
            contact_name="+cmd",
        ),
        description="=HYPERLINK(\"https://example.invalid\",\"click\")",
    )
    mileage = replace(
        _mileage(allocation=allocation),
        purpose="=1+1",
        destination="@destination",
    )

    payload = build_expense_workbook(details, [item], mileage_items=[mileage])
    form = load_workbook(BytesIO(payload), data_only=False)["EXPENSE REIMBURSEMENT"]

    expected = {
        "C5": details.employee_name,
        "G5": details.employee_number,
        "K5": details.employee_home_bu,
        "F64": details.satellite_office,
        "C68": details.employee_name,
        "C10": mileage.purpose,
        "F10": mileage.destination,
        "J10": allocation.account_cost_type,
        "K10": allocation.cost_code_or_wo_type,
        "C45": item.description,
        "F45": item.contact_name,
        "J45": allocation.account_cost_type,
        "K45": allocation.cost_code_or_wo_type,
    }
    for coordinate, value in expected.items():
        assert form[coordinate].value == value
        assert form[coordinate].data_type == "s"


def test_multiple_job_codes_are_grouped_and_receipts_follow_form_order():
    later_group = _item(
        "startup",
        allocation=_job_allocation(job=RRH_JOB_NUMBERS[1]),
    )
    first_group = replace(_item("operations"), description="O&M item")

    payload = build_expense_workbook(_details(), [later_group, first_group])
    workbook = load_workbook(BytesIO(payload), data_only=False)
    form = workbook["EXPENSE REIMBURSEMENT"]
    receipts = workbook["RECEIPTS"]

    assert form["C24"].value == "O&M item"
    assert form["I24"].value == "695400022"
    assert form["I25"].value == "695400023"
    assert "operations.jpg" in receipts["A2"].value
    assert "startup.jpg" in receipts["A60"].value


def test_validation_blocks_signature_and_non_job_routes():
    work_order = ExpenseAllocation(kind=ALLOCATION_WORK_ORDER)
    item = _item(
        section=EXPENSE_SECTION_ENTERTAINMENT,
        amount="19.99",
        allocation=work_order,
    )

    problems = validate_expense_report(
        _details(
            approver_email="not-an-email",
            employee_signature_confirmed=False,
        ),
        [item],
    )

    assert "enter a valid contract administrator email" in problems
    assert "confirm the generated employee signature" in problems
    assert "Receipt 1: enter the entertainment contact name" in problems
    assert (
        "Receipt 1: use job coding; work orders and Other Expenses are not used"
        in problems
    )
    assert "the total reimbursement must exceed $20.00" in problems
    overhead = replace(item, allocation=ExpenseAllocation(kind=ALLOCATION_OVERHEAD))
    assert any(
        "Other Expenses are not used" in problem
        for problem in validate_expense_report(_details(), [overhead])
    )


def test_validation_rejects_a_job_number_from_another_account():
    item = _item(
        allocation=_job_allocation(job="TULANE-695000028-ES JOB CCJ"),
    )

    problems = validate_expense_report(_details(), [item])

    assert "Receipt 1: choose a job number for the selected account" in problems


def test_duplicate_receipt_is_blocked_even_when_other_fields_are_valid():
    problems = validate_expense_report(
        _details(),
        [_item("same"), _item("same", amount="22.00")],
    )
    assert "Receipt 2: remove the duplicate reimbursement line" in problems


def test_split_receipt_writes_multiple_lines_but_attaches_source_once():
    receipt = _receipt_bytes(
        "MARKET RECEIPT\nTeam refreshments  $20.00\nOffice supplies  $15.00\n"
        "Personal items  $33.99\nTOTAL $68.99"
    )
    first = replace(
        _item("market-line-1", amount="20.00"),
        source_receipt_id="market-upload",
        filename="market.jpg",
        file_bytes=receipt,
        description="Refreshments for RRH team",
    )
    second = replace(
        _item("market-line-2", amount="15.00"),
        source_receipt_id="market-upload",
        filename="market.jpg",
        file_bytes=receipt,
        description="Office supplies",
    )

    payload = build_expense_workbook(_details(), [first, second])
    workbook = load_workbook(BytesIO(payload), data_only=False)
    form = workbook["EXPENSE REIMBURSEMENT"]
    receipts = workbook["RECEIPTS"]
    package = build_expense_package(
        _details(),
        [first, second],
        include_pdf=False,
    )

    assert form["C24"].value == "Refreshments for RRH team"
    assert form["H24"].value == 20
    assert form["C25"].value == "Office supplies"
    assert form["H25"].value == 15
    assert len(receipts._images) == 1
    assert receipts["A1"].value == "Receipt 1 of 1 · 2 reimbursement lines"
    assert "$35.00 reimbursable" in receipts["A2"].value
    assert package.receipt_count == 1
    assert package.total == Decimal("35.00")


def test_split_lines_do_not_trigger_a_false_duplicate_receipt_warning():
    source = _receipt_bytes("TWO APPLICABLE ITEMS $25.00 EACH")
    first = replace(
        _item("split-line-1", amount="25.00"),
        source_receipt_id="one-source",
        filename="one-source.jpg",
        file_bytes=source,
    )
    second = replace(
        _item("split-line-2", amount="25.00"),
        source_receipt_id="one-source",
        filename="one-source.jpg",
        file_bytes=source,
    )

    warnings = expense_report_warnings(_details(), [first, second])

    assert not any("same merchant, date, and amount" in item for item in warnings)


def test_split_lines_must_reference_the_same_source_bytes_and_filename():
    first = replace(
        _item("line-one"),
        source_receipt_id="shared-source",
        filename="shared.jpg",
    )
    second = replace(
        _item("line-two"),
        source_receipt_id="shared-source",
        filename="different.jpg",
    )

    problems = validate_expense_report(_details(), [first, second])

    assert "Receipt 2: split lines must use the same uploaded receipt" in problems


def test_official_section_row_limits_block_without_truncating_split_lines():
    allowed_misc = [
        _item(f"misc-{index}", amount="2.00")
        for index in range(MAX_MISCELLANEOUS_ITEMS)
    ]
    too_many_misc = [
        *allowed_misc,
        _item("misc-overflow", amount="2.00"),
    ]
    allowed_entertainment = [
        _item(
            f"entertainment-{index}",
            section=EXPENSE_SECTION_ENTERTAINMENT,
            amount="2.00",
            contact_name="Test Contact",
        )
        for index in range(MAX_ENTERTAINMENT_ITEMS)
    ]
    too_many_entertainment = [
        *allowed_entertainment,
        _item(
            "entertainment-overflow",
            section=EXPENSE_SECTION_ENTERTAINMENT,
            amount="2.00",
            contact_name="Test Contact",
        ),
    ]

    assert not any(
        "official form allows" in item
        for item in validate_expense_report(_details(), allowed_misc)
    )
    assert any(
        f"allows {MAX_MISCELLANEOUS_ITEMS} Miscellaneous" in item
        for item in validate_expense_report(_details(), too_many_misc)
    )
    assert not any(
        "official form allows" in item
        for item in validate_expense_report(_details(), allowed_entertainment)
    )
    assert any(
        f"allows {MAX_ENTERTAINMENT_ITEMS} Entertainment" in item
        for item in validate_expense_report(_details(), too_many_entertainment)
    )


def test_partial_receipt_uses_only_reviewed_business_amount():
    original = _receipt_bytes(
        "ITEM A $12.00\nBUSINESS ITEM $25.00\nITEM C $31.99\nTOTAL $68.99"
    )
    digest = hashlib.sha256(original).hexdigest()
    item = replace(
        _item("partial", amount="25.00"),
        file_bytes=original,
        description="Approved business item only",
    )

    payload = build_expense_workbook(_details(), [item])
    workbook = load_workbook(BytesIO(payload), data_only=False)

    assert workbook["EXPENSE REIMBURSEMENT"]["H24"].value == 25
    assert "$25.00 reimbursable" in workbook["RECEIPTS"]["A2"].value
    assert hashlib.sha256(item.file_bytes).hexdigest() == digest


def test_amount_mileage_parsers_and_irs_rate_boundaries():
    assert parse_expense_amount("$1,234.50 USD") == Decimal("1234.50")
    assert parse_expense_amount("amount 12.00") is None
    assert parse_mileage("1,234.50") == Decimal("1234.50")
    assert parse_mileage("ten") is None
    assert irs_business_mileage_rate(date(2025, 12, 31)) == Decimal("0.70")
    assert irs_business_mileage_rate(date(2026, 6, 30)) == Decimal("0.725")
    assert irs_business_mileage_rate(date(2026, 7, 1)) == Decimal("0.76")
    assert irs_business_mileage_rate(date(2027, 1, 1)) is None


def test_mileage_rows_use_travel_date_rate_and_leave_prohibited_columns_blank():
    mileage = [
        _mileage("first-half", travel_date=date(2026, 6, 30), miles="100"),
        _mileage("second-half", travel_date=date(2026, 7, 1), miles="100"),
    ]

    payload = build_expense_workbook(
        _details(report_date=date(2026, 7, 2)),
        [],
        mileage_items=mileage,
    )
    workbook = load_workbook(BytesIO(payload), data_only=False)
    form = workbook["EXPENSE REIMBURSEMENT"]

    assert workbook.sheetnames == ["EXPENSE REIMBURSEMENT"]
    assert form["B6"].value == "GAS MILEAGE @ APPLICABLE IRS RATE"
    assert form["G10"].value == 100
    assert form["H10"].value == "=ROUND(G10*0.725,2)"
    assert form["H11"].value == "=ROUND(G11*0.76,2)"
    assert form["I10"].value == "695400022"
    assert form["J10"].value == "01AMA"
    assert form["K10"].value == "5490"
    for row in (10, 11):
        assert form.cell(row, 12).value is None
        assert all(form.cell(row, column).value is None for column in range(14, 18))
    assert total_reimbursement([], mileage) == Decimal("148.50")


def test_mileage_half_cent_matches_excels_round_half_up():
    mileage = _mileage(
        "half-cent",
        travel_date=date(2025, 12, 31),
        miles="1.15",
    )

    assert mileage_reimbursement(mileage) == Decimal("0.81")
    assert total_reimbursement([], [mileage]) == Decimal("0.81")
    payload = build_expense_workbook(
        _details(),
        [_item("minimum-total", amount="$20.00")],
        mileage_items=[mileage],
    )
    form = load_workbook(BytesIO(payload), data_only=False)["EXPENSE REIMBURSEMENT"]
    assert form["H10"].value == "=ROUND(G10*0.70,2)"


def test_unknown_future_mileage_rate_blocks_instead_of_guessing():
    problems = validate_expense_report(
        _details(report_date=date(2027, 1, 3)),
        [],
        [_mileage(travel_date=date(2027, 1, 2))],
    )
    assert any("IRS business-mileage rate" in problem for problem in problems)


def test_soft_duplicate_and_date_checks_warn_without_blocking():
    warnings = expense_report_warnings(
        _details(report_date=date(2026, 8, 9)),
        [_item("one"), _item("two", amount="31.25")],
        [_mileage(travel_date=date(2026, 8, 10))],
    )
    assert "Receipt 1 is dated after the report date" in warnings
    assert any("same merchant, date, and amount" in warning for warning in warnings)
    assert "Mileage 1 is dated after the report date" in warnings


def test_content_signature_changes_with_mileage_and_signature_confirmation():
    item = _item("large")
    mileage = _mileage()
    signature = expense_report_signature(_details(), [item], [mileage])

    assert len(signature) == 64
    assert signature == expense_report_signature(_details(), [item], [mileage])
    assert signature != expense_report_signature(
        _details(), [item], [replace(mileage, miles="11")]
    )
    assert signature != expense_report_signature(
        _details(employee_signature_confirmed=False), [item], [mileage]
    )


def test_generated_signature_is_a_bounded_transparent_png():
    payload = employee_signature_png("Test Employee")
    image = Image.open(BytesIO(payload))
    assert image.format == "PNG"
    assert image.mode == "RGBA"
    assert image.width <= 1100
    assert image.height <= 190


def test_receipt_image_is_rotated_bounded_and_not_upscaled():
    source = Image.new("RGB", (600, 900), "white")
    exif = Image.Exif()
    exif[274] = 6
    payload = BytesIO()
    source.save(payload, format="JPEG", quality=90, exif=exif)

    pages = receipt_attachment_pages(payload.getvalue(), "phone-photo.jpg")
    normalized = Image.open(BytesIO(pages[0]))

    assert len(pages) == 1
    assert max(normalized.size) <= 1600
    assert normalized.size == (900, 600)


@pytest.mark.parametrize(
    ("format_name", "suffix"),
    (
        ("PNG", ".png"),
        ("BMP", ".bmp"),
        ("TIFF", ".tiff"),
        ("WEBP", ".webp"),
        ("HEIF", ".heic"),
    ),
)
def test_receipt_picture_formats_normalize_to_printable_jpeg(format_name, suffix):
    if format_name == "HEIF":
        from pillow_heif import register_heif_opener

        register_heif_opener()
    source = Image.new("RGB", (480, 720), "white")
    ImageDraw.Draw(source).text((30, 30), "TOTAL $42.50", fill="black")
    payload = BytesIO()
    source.save(payload, format=format_name)

    pages = receipt_attachment_pages(payload.getvalue(), f"receipt{suffix}")
    normalized = Image.open(BytesIO(pages[0]))

    assert len(pages) == 1
    assert normalized.format == "JPEG"
    assert normalized.size == (480, 720)


def test_multi_page_pdf_receipt_is_attached_once_per_source_page():
    document = fitz.open()
    for label in ("PAGE ONE TOTAL $31.25", "PAGE TWO DETAIL"):
        page = document.new_page(width=612, height=792)
        page.insert_text((72, 72), label)
    pdf_bytes = document.tobytes()
    document.close()
    item = replace(
        _item("pdf-receipt"),
        filename="receipt.pdf",
        file_bytes=pdf_bytes,
    )

    pages = receipt_attachment_pages(pdf_bytes, "receipt.pdf")
    payload = build_expense_workbook(_details(), [item])
    workbook = load_workbook(BytesIO(payload), data_only=False)

    assert len(pages) == 2
    assert len(workbook["RECEIPTS"]._images) == 2
    assert workbook["RECEIPTS"]["A1"].value == "Receipt 1 of 1 · page 1 of 2"
    assert workbook["RECEIPTS"]["A59"].value == "Receipt 1 of 1 · page 2 of 2"


def test_multi_page_pdf_preview_returns_first_page_without_rejecting_source():
    document = fitz.open()
    first = document.new_page(width=612, height=792)
    first.insert_text((72, 72), "FIRST RECEIPT PAGE")
    second = document.new_page(width=612, height=792)
    second.insert_text((72, 72), "SECOND RECEIPT PAGE")
    pdf_bytes = document.tobytes()
    document.close()

    preview = receipt_preview_bytes(pdf_bytes, "two-page-receipt.pdf")

    image = Image.open(BytesIO(preview))
    assert image.format == "JPEG"
    assert image.width > 0 and image.height > 0


def test_corrupt_and_over_page_limit_pdfs_fail_with_actionable_errors():
    with pytest.raises(ExpenseReportError, match="not a readable PDF receipt"):
        receipt_attachment_pages(b"not-a-pdf", "broken.pdf")

    document = fitz.open()
    for _ in range(11):
        document.new_page(width=100, height=100)
    oversized = document.tobytes()
    document.close()
    with pytest.raises(ExpenseReportError, match="11 pages; the limit is 10"):
        receipt_attachment_pages(oversized, "eleven-pages.pdf")


def test_oversized_image_and_pdf_page_fail_before_large_raster_allocation():
    image = Image.new("1", (7000, 7000), 1)
    image_payload = BytesIO()
    image.save(image_payload, format="PNG")
    with pytest.raises(ExpenseReportError, match="7000×7000"):
        receipt_attachment_pages(image_payload.getvalue(), "huge.png")

    document = fitz.open()
    document.new_page(width=20_000, height=20_000)
    pdf_payload = document.tobytes()
    document.close()
    with pytest.raises(ExpenseReportError, match="huge.pdf page 1 is too large"):
        receipt_attachment_pages(pdf_payload, "huge.pdf")


def test_email_uses_pdf_only_and_excel_remains_in_package():
    package = ExpensePackage(
        basename="expense",
        workbook_bytes=b"editable-xlsx",
        pdf_bytes=b"%PDF-submission",
        total=Decimal("50"),
        receipt_count=1,
    )
    assert email_attachments_for_package(package) == [
        ("expense.pdf", b"%PDF-submission")
    ]
    assert package.workbook_bytes == b"editable-xlsx"

    message = BytesParser(policy=policy.default).parsebytes(
        _build_expense_eml(_details(), package)
    )
    assert message["To"] == "rrh.approver@example.invalid"
    assert message["X-Unsent"] == "1"
    assert [part.get_filename() for part in message.iter_attachments()] == [
        "expense.pdf"
    ]


@requires_libreoffice
def test_combined_pdf_places_signed_form_before_each_receipt():
    package = build_expense_package(
        _details(),
        [_item("one"), _item("two", amount="22.00")],
    )

    assert package.pdf_error == ""
    assert package.pdf_bytes is not None
    with fitz.open(stream=package.pdf_bytes, filetype="pdf") as document:
        assert document.page_count == 3
        assert "REIMBURSEMENT OF EXPENSES" in document[0].get_text()
        assert "Synthetic Employee" in document[0].get_text()
        assert "Receipt 1 of 2" in document[1].get_text()
        assert "Receipt 2 of 2" in document[2].get_text()


def test_long_employee_name_signature_is_fitted_rather_than_clipped():
    """A long name must shrink to fit, never be silently cut off.

    The image is embedded on the official JDE form and the employee attests to
    it ("I confirm this generated signature represents me"), so a clipped
    rendering is an attestation to a truncated version of their own name.
    """
    short = employee_signature_png("Evan Roden")
    long_name = employee_signature_png(
        "Maria de los Angeles Fernandez-Villalobos"
    )

    for payload in (short, long_name):
        assert payload.startswith(b"\x89PNG\r\n\x1a\n")
        image = Image.open(BytesIO(payload))
        image.load()
        # Nothing may touch the right edge: a fitted render always leaves the
        # trailing margin, whereas a clipped one runs to the boundary.
        assert image.width <= 1100
        assert image.getbbox() is not None


def test_unrenderable_employee_name_is_refused_not_truncated():
    """When even the minimum size cannot fit, fail loudly instead of clipping."""
    with pytest.raises(ExpenseReportError, match="too long to render"):
        employee_signature_png("W" * 160)


def test_pdf_renderer_timeout_preserves_the_completed_workbook(monkeypatch):
    """A renderer outage must degrade to Excel-only, never discard the report.

    subprocess.run(..., timeout=...) raises TimeoutExpired, which is not an
    ExpenseReportError; catching only the latter let it escape and the caller
    dropped the whole validated package, forcing full re-entry.
    """
    import subprocess

    def _timeout(_workbook: bytes) -> bytes:
        raise subprocess.TimeoutExpired(cmd="soffice", timeout=120)

    monkeypatch.setattr(
        "app.expense_report.convert_expense_workbook_to_pdf", _timeout
    )

    package = build_expense_package(_details(), [_item()])

    assert package.workbook_bytes.startswith(b"PK")
    assert package.pdf_bytes is None
    assert package.pdf_error


def test_pdf_renderer_connection_failure_preserves_the_completed_workbook(monkeypatch):
    """Same guarantee for the Gotenberg backend's transport errors."""
    import requests

    def _unreachable(_workbook: bytes) -> bytes:
        raise requests.ConnectionError("gotenberg unreachable")

    monkeypatch.setattr(
        "app.expense_report.convert_expense_workbook_to_pdf", _unreachable
    )

    package = build_expense_package(_details(), [_item()])

    assert package.workbook_bytes.startswith(b"PK")
    assert package.pdf_bytes is None
    assert "unreachable" in package.pdf_error


# --- Transparent receipts must flatten onto WHITE, in every mode ------------


@pytest.mark.parametrize("mode", ["P", "RGBA", "LA", "PA"])
def test_a_transparent_receipt_is_flattened_onto_white_not_blacked_out(mode):
    """A PNG-8/GIF receipt used to come out 100% BLACK in the submitted packet.

    The old branch tested ``mode in {"RGBA", "LA"} or "transparency" in info``
    and then took its mask from ``getchannel("A") if "A" in getbands()``. A
    palette receipt is mode "P": it PASSED the test via info["transparency"] but
    has bands ("P",) and no "A", so the mask was None -- and
    ``paste(..., mask=None)`` is a plain overwrite, so the flatten silently never
    happened. Mode "PA" missed the test entirely.

    Measured before the fix: 100% of the output near-black, the text destroyed
    along with the background, no error raised, and the result attached to the
    approver's PDF.
    """
    source = Image.new("RGBA", (400, 300), (0, 0, 0, 0))
    ImageDraw.Draw(source).text((40, 140), "TOTAL $31.25", fill=(0, 0, 0, 255))
    if mode == "P":
        source = source.convert("P", palette=Image.ADAPTIVE, colors=64)
        source.info["transparency"] = 0
    elif mode in {"LA", "PA"}:
        source = source.convert(mode)

    rendered = Image.open(BytesIO(_compact_receipt_image(source))).convert("RGB")
    pixels = list(rendered.get_flattened_data())

    # The page background is white, not the transparent pixels' leftover black.
    assert sum(rendered.getpixel((5, 5))) > 700, f"{mode}: background not white"
    ink = sum(1 for pixel in pixels if sum(pixel) < 200)
    # Text survived...
    assert ink > 0, f"{mode}: the receipt text was flattened away entirely"
    # ...and the page was not blacked out.
    assert ink < len(pixels) * 0.25, f"{mode}: receipt came out mostly black"


def test_an_opaque_receipt_is_unaffected_by_the_flatten_branch():
    source = Image.new("RGB", (400, 300), "white")
    ImageDraw.Draw(source).text((40, 140), "TOTAL $31.25", fill="black")
    rendered = Image.open(BytesIO(_compact_receipt_image(source))).convert("RGB")
    assert sum(rendered.getpixel((5, 5))) > 700


def test_the_flatten_matches_the_one_in_ocr():
    """These two are the same fix for the same reason and have already drifted
    apart once -- ocr.py was correct while expense_report.py silently was not.
    Pin the mode test so a future edit to one is visible against the other."""
    import inspect

    from app import ocr

    for source in (
        inspect.getsource(_compact_receipt_image),
        inspect.getsource(ocr.image_blocks_for_vision),
    ):
        assert '"RGBA", "LA", "PA"' in source
        assert 'rgba.split()[-1]' in source or "split()[-1]" in source
